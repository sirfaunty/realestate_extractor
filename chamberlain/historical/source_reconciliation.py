"""Multi-year MRI vs CBL/Village Green reconciliation.

For each year where both an MRI export and a CBL property-manager
statement exist (2023-2025), pull the key rollups from each and
explain the difference structurally:

  1. TIF receivable: MRI books it as income; CBL excludes it.
  2. Major Expense (routine capital): CBL places below NOI;
     MRI rolls into operating expenses (above NOI).
  3. Residual: minor line-level reclass + RE-tax timing.

The CBL "property-operations" NOI (ex-TIF, ex-capital) is the basis
the model carries, matching its architecture (separate TIF sub-engine
+ separate CapEx schedule).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
import pdfplumber


@dataclass
class YearRecon:
    year: int
    mri_total_income: Optional[float] = None
    mri_total_opex: Optional[float] = None
    mri_noi: Optional[float] = None
    mri_tif_income: Optional[float] = None
    cbl_total_income: Optional[float] = None
    cbl_total_opex: Optional[float] = None
    cbl_noi: Optional[float] = None
    cbl_major_expense: Optional[float] = None
    cbl_tif_income: Optional[float] = None
    notes: list[str] = field(default_factory=list)

    @property
    def income_delta(self) -> float:
        return (self.mri_total_income or 0) - (self.cbl_total_income or 0)

    @property
    def noi_delta(self) -> float:
        return (self.mri_noi or 0) - (self.cbl_noi or 0)

    @property
    def tif_explains_income(self) -> float:
        """Income delta after netting TIF on BOTH sides.

        If CBL also booked TIF (FY2023), TIF cancels and the residual is
        just the rental/other-income reclass. If only MRI has TIF
        (FY2024-25), removing MRI TIF brings the two onto a comparable
        ex-TIF basis.
        """
        mri_ex = (self.mri_total_income or 0) - (self.mri_tif_income or 0)
        cbl_ex = (self.cbl_total_income or 0) - (self.cbl_tif_income or 0)
        return mri_ex - cbl_ex


_NUM = re.compile(r"-?[\d,]+\.\d{2}")


def _cbl_rollups(pdf_path: Path) -> dict:
    """Extract CBL statement rollups (12-month Total = last number).

    Also detects whether CBL booked the TIF receivable itself (account
    4961-100 MISC INCOME - TIF REC). In some years (FY2023) CBL includes
    TIF in income just like MRI; in others (FY2024 zeroed, FY2025 absent)
    it does not. The reconciliation must know this to avoid spuriously
    attributing the whole gap to TIF.
    """
    out: dict[str, float] = {}
    with pdfplumber.open(pdf_path) as pdf:
        full = "\n".join((p.extract_text() or "") for p in pdf.pages)
    targets = {
        "TOTAL INCOME": "total_income",
        "TOTAL OPERATING EXPENSES": "total_opex",
        "NET OPERATING INCOME": "noi",
        "TOTAL MAJOR EXPENSE": "major_expense",
    }
    for line in full.split("\n"):
        up = line.upper().strip()
        # CBL TIF account line: "4961-100 MISC INCOME - TIF REC ... <total>"
        if "TIF" in up and ("4961-100" in up or "MISC INCOME - TIF" in up):
            nums = _NUM.findall(line)
            if nums and "cbl_tif_income" not in out:
                out["cbl_tif_income"] = float(nums[-1].replace(",", ""))
        for label, key in targets.items():
            if key in out:
                continue
            if not up.startswith(label):
                continue
            rest = up[len(label):].lstrip()
            if rest and not (rest[0].isdigit() or rest[0] == "-"):
                continue
            nums = _NUM.findall(line)
            if len(nums) >= 13:
                out[key] = float(nums[-1].replace(",", ""))
    return out


def _mri_rollups(xlsx_path: Path) -> dict:
    """Extract MRI statement rollups + TIF income line.

    The standard MRI 12-month export has 12 month columns + a single
    'Total' column. Reforecast exports add 'Total Budgeted' and
    'Variance' columns after 'Total Forecast'. We must pick the actuals/
    forecast Total column, NOT Budgeted or Variance. Strategy: scan the
    header row for a cell whose text starts with 'Total' and take the
    FIRST such column (Total / Total Forecast); ignore later
    Budgeted/Variance columns.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # Locate header row (the one carrying month labels like 2024-01)
    total_col = ws.max_column
    for hr in range(1, min(14, ws.max_row + 1)):
        row_txt = {
            c: str(ws.cell(row=hr, column=c).value).strip().lower()
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=hr, column=c).value is not None
        }
        # the header row also has a 'total' cell
        total_cols = [c for c, t in row_txt.items() if t.startswith("total")]
        has_months = any(
            isinstance(ws.cell(row=hr, column=c).value, str)
            and re.match(r"\d{4}-\d{2}", str(ws.cell(row=hr, column=c).value))
            for c in range(1, ws.max_column + 1)
        )
        if total_cols and (has_months or hr >= 8):
            total_col = min(total_cols)  # first Total = actuals/forecast
            break

    out: dict[str, float] = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if not label:
            continue
        s = str(label).strip().upper()
        v = ws.cell(row=r, column=total_col).value
        if not isinstance(v, (int, float)):
            continue
        if s == "TOTAL INCOME" and "total_income" not in out:
            out["total_income"] = float(v)
        elif s == "TOTAL OPERATING EXPENSES" and "total_opex" not in out:
            out["total_opex"] = float(v)
        elif s == "NET OPERATING INCOME" and "noi" not in out:
            out["noi"] = float(v)
        elif "TIF" in s and "RECEIV" in s and "tif_income" not in out:
            out["tif_income"] = float(v)
    wb.close()
    return out


# Year -> (MRI xlsx filename, CBL pdf filename)
_YEAR_SOURCES = {
    2023: ("Chamb 2023 Actuals.XLSX", "CBL 2023 12 Month Statement.pdf"),
    2024: ("Chamb 2024 Reforecast.XLSX", "CBL 2024 12 Month Statement.pdf"),
    2025: ("Chamberlain MRI 2025 12 Month Income Statement.xlsx",
            "CBL 2025 12 Month Statement.pdf"),
}


def reconcile_mri_vs_cbl(
    mri_actuals_dir: Path,
) -> list[YearRecon]:
    """Build the MRI-vs-CBL reconciliation for all overlapping years."""
    results: list[YearRecon] = []
    for year, (mri_fn, cbl_fn) in sorted(_YEAR_SOURCES.items()):
        mri_path = mri_actuals_dir / mri_fn
        cbl_path = mri_actuals_dir / cbl_fn
        rec = YearRecon(year=year)

        if mri_path.exists():
            m = _mri_rollups(mri_path)
            rec.mri_total_income = m.get("total_income")
            rec.mri_total_opex = m.get("total_opex")
            rec.mri_noi = m.get("noi")
            rec.mri_tif_income = m.get("tif_income")
        if cbl_path.exists():
            c = _cbl_rollups(cbl_path)
            rec.cbl_total_income = c.get("total_income")
            rec.cbl_total_opex = c.get("total_opex")
            rec.cbl_noi = c.get("noi")
            rec.cbl_major_expense = c.get("major_expense")
            rec.cbl_tif_income = c.get("cbl_tif_income")

        # Structural explanation
        mri_tif = rec.mri_tif_income or 0
        cbl_tif = rec.cbl_tif_income or 0
        if mri_tif and cbl_tif:
            rec.notes.append(
                f"TIF receivable ${mri_tif:,.0f} is booked in BOTH systems "
                f"(MRI Misc Revenue-TIF; CBL acct 4961-100) — it cancels in "
                f"the bridge."
            )
        elif mri_tif and not cbl_tif:
            rec.notes.append(
                f"MRI income includes TIF receivable ${mri_tif:,.0f}; "
                f"CBL excludes it (account zeroed/absent this year)."
            )
        if rec.cbl_major_expense:
            rec.notes.append(
                f"CBL places Major Expense (routine capital) "
                f"${rec.cbl_major_expense:,.0f} below NOI; MRI rolls it "
                f"into operating expenses (above NOI)."
            )
        if rec.mri_total_income and rec.cbl_total_income:
            resid = rec.tif_explains_income
            rec.notes.append(
                f"Income reconciles to ${abs(resid):,.0f} residual on an "
                f"ex-TIF basis (rental vs other-income reclass). NOI bridge: "
                f"income gap ${rec.income_delta:+,.0f} less opex gap "
                f"${(rec.mri_total_opex or 0) - (rec.cbl_total_opex or 0):+,.0f}"
                f" = ${rec.noi_delta:+,.0f}."
            )
        results.append(rec)
    return results


def format_recon_report(recs: list[YearRecon]) -> str:
    """Human-readable multi-year reconciliation report."""
    out: list[str] = []
    out.append("MRI (KARE) vs CBL (Village Green) — multi-year reconciliation")
    out.append("=" * 64)
    for r in recs:
        out.append(f"\nFY{r.year}")
        out.append(f"  {'':22}{'MRI':>15}{'CBL':>15}{'Delta':>14}")
        rows = [
            ("Total Income", r.mri_total_income, r.cbl_total_income),
            ("Total OpEx", r.mri_total_opex, r.cbl_total_opex),
            ("NOI", r.mri_noi, r.cbl_noi),
        ]
        for lbl, m, c in rows:
            m = m or 0
            c = c or 0
            out.append(f"  {lbl:22}{m:>15,.0f}{c:>15,.0f}{m - c:>14,.0f}")
        if r.mri_tif_income:
            out.append(f"  MRI TIF receivable    {r.mri_tif_income:>15,.0f}")
            out.append(f"  MRI NOI ex-TIF        "
                       f"{(r.mri_noi or 0) - r.mri_tif_income:>15,.0f}")
        for n in r.notes:
            out.append(f"  - {n}")
    return "\n".join(out)


__all__ = ["YearRecon", "reconcile_mri_vs_cbl", "format_recon_report"]
