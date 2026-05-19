"""MRI 12-Month Income Statement loader.

Parses the standardized MRI export format used across Chamb {YYYY} Actuals.XLSX
files and the TTM Sep 2025 file. Output is one HistoricalPeriod per year
populated with line items + roll-up subtotals (where present in source).

MRI export structure observed in the corpus:
  - Single sheet, typically 'MRI_12MINCS' or 'KA_FORECAST' (2024 reforecast).
  - Header rows 1-6: Database, Report Id, Kraus-Anderson, blank, "Through Period",
    Accrual.
  - Row 9: column headers (12 monthly dates + 'Total' or 'Forecast').
  - Rows 10+: P&L line items. Column A has the label, columns B-M have
    monthly amounts, column N has the year total. (2024 reforecast file
    has an extra forecast column.)
  - Subtotal rows are styled but indistinguishable from line items by
    cell content alone — we identify them via the SUBTOTAL_LABELS set.
  - Header-only category rows (REVENUE, RENTAL REVENUE-MULTI-FAMILY, etc.)
    have no values in the data columns; they're recognized via SUBTOTAL_LABELS.

The loader produces:
  - One HistoricalPeriod per year (period_type=YEAR)
  - HistoricalLineItem per non-subtotal, non-skip row with a non-zero Total
  - Subtotal rows (REVENUE, TOTAL OPERATING EXPENSES, NOI, etc.) captured
    on the HistoricalPeriod's rollup fields (total_revenue, total_opex,
    noi_as_reported) when found in source

Every line item carries a Citation pointing to:
  - Source document id (e.g. 'mri_chamb_2023_actuals')
  - Locator: sheet name + cell address of the Total column for that row

This is the high-water mark for citation density in the model — every
historical line item has a clickable drill-back.
"""

from __future__ import annotations

import logging
import re
from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl

from .line_item_mapper import classify_label
from ..models.citation import (
    Citation,
    Cited,
    Locator,
    SourceDocumentRegistry,
)
from ..models.historical import (
    FiscalPeriod,
    FiscalPeriodType,
    HistoricalActuals,
    HistoricalLineCategory,
    HistoricalLineItem,
    HistoricalPeriod,
)

logger = logging.getLogger(__name__)


# Column letter sequence for openpyxl column-index → letter conversion.
def _col_letter(idx: int) -> str:
    """1-indexed column → Excel letter."""
    letters = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


# Subtotal labels that map to HistoricalPeriod rollup fields rather than
# individual HistoricalLineItem records.
ROLLUP_TARGETS: dict[str, str] = {
    "TOTAL INCOME": "total_revenue",
    "TOTAL OPERATING EXPENSES": "total_opex_as_reported",
    "NET OPERATING INCOME": "noi_as_reported",
}


def _find_header_row(ws) -> tuple[int, list[Optional[date]]]:
    """Locate the row containing the 12 monthly date headers.

    Returns (header_row_index, list_of_date_per_column).
    """
    # Scan rows 5-12 for the first row where column B is a date or year string.
    for r in range(5, 15):
        b_val = ws.cell(row=r, column=2).value
        if b_val is None:
            continue
        if hasattr(b_val, "year"):
            # It's a datetime; build the date list
            dates: list[Optional[date]] = []
            for c in range(2, 14):
                v = ws.cell(row=r, column=c).value
                if hasattr(v, "year"):
                    dates.append(date(v.year, v.month, 1))
                else:
                    dates.append(None)
            return r, dates
    raise ValueError("Could not locate monthly header row in MRI sheet")


def _total_column_index(ws, header_row: int) -> int:
    """Find the Total / Forecast column.

    Usually column N (14) but the 2024 reforecast file has 'Total Forecast'
    further to the right. Strategy: find the rightmost column in the header
    row whose value is a string containing 'Total' or 'Forecast'.
    """
    for c in range(ws.max_column, 13, -1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and ("Total" in v or "Forecast" in v):
            return c
    # Default: column N
    return 14


def load_mri_file(
    file_path: Path,
    source_document_id: str,
    fiscal_year: int,
    registry: SourceDocumentRegistry,
) -> HistoricalPeriod:
    """Parse one MRI 12-Month Income Statement file into a HistoricalPeriod.

    Args:
        file_path: path to the .XLSX file
        source_document_id: registry id of the SourceDocument
        fiscal_year: the fiscal year being loaded
        registry: SourceDocumentRegistry (used to verify source_document_id exists)

    Returns:
        HistoricalPeriod with line items + rollups populated.
    """
    registry.require(source_document_id)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    sheet_name = ws.title

    header_row, _monthly_dates = _find_header_row(ws)
    total_col = _total_column_index(ws, header_row)
    total_col_letter = _col_letter(total_col)

    period = HistoricalPeriod(
        period=FiscalPeriod(
            period_type=FiscalPeriodType.YEAR,
            year=fiscal_year,
            label=f"{fiscal_year}A",
        ),
    )

    line_items: list[HistoricalLineItem] = []
    unknown_labels: list[str] = []

    for r in range(header_row + 1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if label is None:
            continue
        label_str = str(label).strip()
        if not label_str:
            continue

        total_val = ws.cell(row=r, column=total_col).value
        if not isinstance(total_val, (int, float)):
            # Header row or empty — skip
            continue

        # Classify the label
        category, method = classify_label(label_str)

        # Capture rollup totals on the HistoricalPeriod
        if method == "subtotal" and label_str in ROLLUP_TARGETS:
            field = ROLLUP_TARGETS[label_str]
            cited_total = Cited(
                value=float(total_val),
                citations=[Citation(
                    source_document_id=source_document_id,
                    locator=Locator(
                        sheet=sheet_name,
                        cell=f"{total_col_letter}{r}",
                        row=r,
                        row_label=label_str,
                    ),
                    verbatim_text=f"{total_val}",
                )],
            )
            setattr(period, field, cited_total)
            continue

        # Subtotal/skip — don't create a line item
        if method in ("subtotal", "skip"):
            continue

        # Track unknown labels
        if method == "unknown":
            unknown_labels.append(label_str)

        # Create a HistoricalLineItem
        # Sign convention: revenue is positive in MRI; expenses are positive
        # in MRI (entered as positive amounts in expense rows). For the
        # canonical line item amount, we keep the MRI sign and let downstream
        # code interpret based on category. Many expense rows are stored as
        # positive in MRI even though they reduce NOI.
        cited_amount = Cited(
            value=float(total_val),
            citations=[Citation(
                source_document_id=source_document_id,
                locator=Locator(
                    sheet=sheet_name,
                    cell=f"{total_col_letter}{r}",
                    row=r,
                    row_label=label_str,
                    col_label="Total",
                ),
                verbatim_text=f"{total_val}",
            )],
        )

        if category is None:
            # Defensive: shouldn't reach here given the subtotal/skip handling above
            continue

        line_items.append(HistoricalLineItem(
            label=label_str,
            category=category,
            amount=cited_amount,
            period=period.period,
        ))

    wb.close()

    period.line_items = line_items

    if unknown_labels:
        logger.warning(
            f"MRI loader: {fiscal_year} had {len(unknown_labels)} unknown labels: "
            f"{unknown_labels[:5]}{'...' if len(unknown_labels) > 5 else ''}"
        )

    return period


def load_all_chamberlain_actuals(
    registry: SourceDocumentRegistry,
    source_docs_root: Optional[Path] = None,
) -> HistoricalActuals:
    """Load all available MRI actuals files (2017-2024 actuals + 2024 reforecast + TTM Sep 2025).

    Uses the SourceDocumentRegistry to map year → document id → file path.

    Returns:
        HistoricalActuals with one HistoricalPeriod per loaded year.
    """
    actuals = HistoricalActuals()

    # 2017-2023 actuals files
    for yr in range(2017, 2024):
        doc_id = f"mri_chamb_{yr}_actuals"
        doc = registry.get(doc_id)
        if doc is None or doc.file_path is None or not doc.file_path.exists():
            logger.info(f"MRI file for {yr} not present; skipping")
            continue
        period = load_mri_file(doc.file_path, doc_id, yr, registry)
        actuals.periods.append(period)
        logger.info(f"Loaded {yr}: {len(period.line_items)} line items, "
                    f"NOI={period.noi_as_reported.value if period.noi_as_reported else 'n/a'}")

    # 2024 reforecast (treat as 2024 actuals — it's the latest 2024 snapshot)
    doc_id = "mri_chamb_2024_reforecast"
    doc = registry.get(doc_id)
    if doc and doc.file_path and doc.file_path.exists():
        period = load_mri_file(doc.file_path, doc_id, 2024, registry)
        # Mark this period as reforecast in the label
        period.period = FiscalPeriod(
            period_type=FiscalPeriodType.YEAR,
            year=2024,
            label="2024 Reforecast",
        )
        actuals.periods.append(period)
        logger.info(f"Loaded 2024 Reforecast: {len(period.line_items)} line items")

    # TTM Sep 2025
    doc_id = "mri_chamb_ttm_sep_2025"
    doc = registry.get(doc_id)
    if doc and doc.file_path and doc.file_path.exists():
        period = load_mri_file(doc.file_path, doc_id, 2025, registry)
        period.period = FiscalPeriod(
            period_type=FiscalPeriodType.TTM,
            year=2025,
            month=9,
            label="TTM Sep 2025",
        )
        actuals.periods.append(period)
        logger.info(f"Loaded TTM Sep 2025: {len(period.line_items)} line items")

    # Calendar-year 2025 actuals (Jan-Sep from TTM Sep + Oct-Dec from CBL).
    # This is the full-year 2025 column for the timeline.
    try:
        cy2025 = load_cy2025_actuals(registry)
        actuals.periods.append(cy2025)
        logger.info(
            "Loaded CY2025 actuals: NOI "
            f"${cy2025.noi_as_reported.value:,.0f}"
            if cy2025.noi_as_reported else "Loaded CY2025 actuals"
        )
    except Exception as exc:  # pragma: no cover - defensive
        logger.warning("Could not build CY2025 actuals: %s", exc)

    return actuals


def load_cbl_ttm_statement(
    registry: SourceDocumentRegistry,
    source_document_id: str = "cbl_ttm_mar_2026",
) -> HistoricalPeriod:
    """Parse the CBL trailing-twelve statement (Apr 2025 - Mar 2026).

    The CBL/Village Green format differs from the MRI export:
      - Row 5 has month headers; data starts row 6
      - Column A = GL code, Column B = line label
      - Columns C..N = 12 monthly values, Column O = Total
      - Subtotal rows have a blank GL code (col A) and an all-caps label

    Returns a HistoricalPeriod labeled "TTM Mar 2026" (period_type=TTM)
    carrying line items + the income/opex/NOI rollups, all cited.
    """
    doc = registry.require(source_document_id)
    assert doc.file_path is not None

    wb = openpyxl.load_workbook(doc.file_path, data_only=True)
    ws = wb.active
    sheet_name = ws.title
    total_col = 15  # column O
    total_letter = "O"

    period = HistoricalPeriod(
        period=FiscalPeriod(
            period_type=FiscalPeriodType.TTM,
            year=2026,
            month=3,
            label="TTM Mar 2026",
        ),
    )

    line_items: list[HistoricalLineItem] = []
    rollup_targets = {
        "TOTAL INCOME": "total_revenue",
        "TOTAL OPERATING EXPENSE": "total_opex_as_reported",
        "TOTAL EXPENSE": "total_opex_as_reported",
        "NET OPERATING INCOME": "noi_as_reported",
    }

    for r in range(6, ws.max_row + 1):
        gl_code = ws.cell(row=r, column=1).value
        label = ws.cell(row=r, column=2).value
        if label is None:
            continue
        label_str = str(label).strip()
        if not label_str:
            continue
        total_val = ws.cell(row=r, column=total_col).value
        if not isinstance(total_val, (int, float)):
            continue

        up = label_str.upper()
        # Rollup subtotal rows have no GL code
        is_subtotal = gl_code is None or str(gl_code).strip() == ""

        if is_subtotal and up in rollup_targets:
            field = rollup_targets[up]
            setattr(period, field, Cited(
                value=float(total_val),
                citations=[Citation(
                    source_document_id=source_document_id,
                    locator=Locator(sheet=sheet_name, cell=f"{total_letter}{r}",
                                    row=r, row_label=label_str),
                    verbatim_text=f"{total_val}",
                )],
            ))
            continue

        if is_subtotal:
            # Other subtotal lines (GROSS POTENTIAL, TOTAL RENTAL INCOME,
            # TOTAL OTHER INCOME, etc.) — skip as line items to avoid
            # double counting; they're recoverable from components.
            continue

        category, _method = classify_label(label_str)
        if category is None:
            continue

        line_items.append(HistoricalLineItem(
            label=label_str,
            category=category,
            amount=Cited(
                value=float(total_val),
                citations=[Citation(
                    source_document_id=source_document_id,
                    locator=Locator(sheet=sheet_name, cell=f"{total_letter}{r}",
                                    row=r, row_label=label_str, col_label="Total"),
                    verbatim_text=f"{total_val}",
                )],
            ),
            period=period.period,
        ))

    wb.close()
    period.line_items = line_items
    return period


def load_cy2025_actuals(
    registry: SourceDocumentRegistry,
) -> HistoricalPeriod:
    """Load full-year 2025 actuals from the CBL 2025 12-Month Statement PDF.

    Single authoritative source: "CBL 2025 12 Month Statement.pdf",
    Period = Jan 2025-Dec 2025, Book = Accrual. The statement carries 12
    monthly columns plus a Total column; we read the Total column for the
    rollup lines (Total Income / Total Operating Expenses / NOI). Every
    figure is cited to that statement.

    Returns a HistoricalPeriod labeled "2025A" (period_type=YEAR).
    """
    import pdfplumber

    doc_id = "cbl_2025_statement"
    doc = registry.require(doc_id)
    assert doc.file_path is not None and doc.file_path.exists(), (
        f"CBL 2025 statement not found at {doc.file_path}"
    )

    period = HistoricalPeriod(
        period=FiscalPeriod(
            period_type=FiscalPeriodType.YEAR,
            year=2025,
            label="2025A",
        ),
    )

    # Map the statement's rollup row labels -> HistoricalPeriod fields
    rollup_targets = {
        "TOTAL INCOME": "total_revenue",
        "TOTAL OPERATING EXPENSES": "total_opex_as_reported",
        "NET OPERATING INCOME": "noi_as_reported",
    }
    found: dict[str, tuple[float, int]] = {}

    num_re = re.compile(r"-?[\d,]+\.\d{2}")
    with pdfplumber.open(doc.file_path) as pdf:
        for pageno, page in enumerate(pdf.pages, start=1):
            for line in (page.extract_text() or "").split("\n"):
                up = line.upper().strip()
                for label, field in rollup_targets.items():
                    if field in found:
                        continue
                    # Exact label match: the line must begin with the label
                    # followed by a digit/space/minus (NOT another letter, so
                    # "TOTAL INCOME ADJUSTMENTS" never matches "TOTAL INCOME").
                    if not up.startswith(label):
                        continue
                    rest = up[len(label):].lstrip()
                    if rest and not (rest[0].isdigit() or rest[0] == "-"):
                        continue
                    nums = num_re.findall(line)
                    # 12 monthly values + 1 Total = 13; Total is last
                    if len(nums) >= 13:
                        total = float(nums[-1].replace(",", ""))
                        found[field] = (total, pageno)

    for field, (val, pageno) in found.items():
        setattr(period, field, Cited(
            value=val,
            citations=[Citation(
                source_document_id=doc_id,
                locator=Locator(
                    page=pageno,
                    note=f"{field} — Total column (Jan-Dec 2025)",
                ),
                verbatim_text=f"{val:,.2f}",
            )],
        ))

    return period


__all__ = ["load_mri_file", "load_all_chamberlain_actuals",
           "load_cbl_ttm_statement", "load_cy2025_actuals"]
