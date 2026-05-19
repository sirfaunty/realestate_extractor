"""Equity ledger loader.

Parses the KA + IDP capital-account GL detail from the Equity Account
Details workbook into a typed, cited EquityLedger.

Source structure (Copy of Chamberlain Equity Account Details.xlsx):
  - 'Equity Summary'        — annual KA/IDP contribution rollup (thin)
  - 'MR29105003 - KA'       — KA capital account GL detail (Period, Entry
                              Date, Ref, Description, Debit, Credit)
  - 'MR29115003 - IDP'      — IDP capital account GL detail (same shape)
  - 'Draws'                 — KA construction-period draw schedule
  - ' Overall Sources & Uses' — development budget (uses side)
  - '2018 Closing Statement'  — inception closing (empty in this copy)

GL sign convention in the source:
  - Credit to capital account = contribution (cash in from investor)
  - Debit to capital account  = distribution (cash out to investor)

Our EquityTransaction.amount convention:
  - contribution  -> NEGATIVE (cash from investor / into LLC)
  - distribution  -> POSITIVE (cash to investor / out of LLC)

So: amount = debit - credit  (credit contribution becomes negative;
debit distribution becomes positive).

Each transaction is cited to the workbook + sheet + row.
"""

from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl

from ..models.citation import (
    Citation,
    Cited,
    Locator,
    SourceDocumentRegistry,
)
from ..models.historical import EquityLedger, EquityTransaction

logger = logging.getLogger(__name__)

# GL sheet name -> investor class id
_ACCOUNT_SHEETS = {
    "MR29105003 - KA": "KA",
    "MR29115003 - IDP": "IDP",
}


def _to_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    # Some entry dates are strings like '2019-08-01 17:14:0'
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:19], fmt).date()
        except ValueError:
            continue
    return None


def _period_to_date(period: str) -> Optional[date]:
    """Parse an MM/YY period string to a date (first of month)."""
    if not period:
        return None
    s = str(period).strip()
    if "/" in s:
        mm, yy = s.split("/")[:2]
        try:
            m = int(mm)
            y = int(yy)
            y = 2000 + y if y < 80 else 1900 + y
            return date(y, m, 1)
        except ValueError:
            return None
    return None


def load_equity_ledger(
    registry: SourceDocumentRegistry,
    source_document_id: str = "equity_account_details",
) -> EquityLedger:
    """Parse the KA + IDP capital-account GL detail into an EquityLedger.

    Args:
        registry: SourceDocumentRegistry (must contain source_document_id)
        source_document_id: id of the Equity Account Details workbook

    Returns:
        EquityLedger with one EquityTransaction per GL entry, cited.
    """
    doc = registry.require(source_document_id)
    assert doc.file_path is not None, "equity_account_details has no file_path"

    wb = openpyxl.load_workbook(doc.file_path, data_only=True)
    ledger = EquityLedger()

    for sheet_name, class_id in _ACCOUNT_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            logger.warning("Equity sheet missing: %s", sheet_name)
            continue
        ws = wb[sheet_name]

        # Locate header row: the row containing 'Period' in col B or A and
        # 'Debit'/'Credit' somewhere to the right.
        header_row = None
        col_map: dict[str, int] = {}
        for r in range(1, min(12, ws.max_row + 1)):
            row_vals = {
                (str(ws.cell(row=r, column=c).value).strip().lower()
                 if ws.cell(row=r, column=c).value is not None else ""): c
                for c in range(1, ws.max_column + 1)
            }
            if "period" in row_vals and ("debit" in row_vals or "credit" in row_vals):
                header_row = r
                for key in ("period", "entry date", "ref", "description",
                            "debit", "credit"):
                    if key in row_vals:
                        col_map[key] = row_vals[key]
                break

        if header_row is None:
            logger.warning("Could not find header in %s", sheet_name)
            continue

        c_period = col_map.get("period")
        c_entry = col_map.get("entry date")
        c_ref = col_map.get("ref")
        c_desc = col_map.get("description")
        c_debit = col_map.get("debit")
        c_credit = col_map.get("credit")

        for r in range(header_row + 1, ws.max_row + 1):
            debit = ws.cell(row=r, column=c_debit).value if c_debit else None
            credit = ws.cell(row=r, column=c_credit).value if c_credit else None
            if not isinstance(debit, (int, float)):
                debit = 0.0
            if not isinstance(credit, (int, float)):
                credit = 0.0
            if debit == 0.0 and credit == 0.0:
                continue

            period = ws.cell(row=r, column=c_period).value if c_period else None
            entry_dt = ws.cell(row=r, column=c_entry).value if c_entry else None
            ref = ws.cell(row=r, column=c_ref).value if c_ref else None
            desc = ws.cell(row=r, column=c_desc).value if c_desc else None
            desc_str = str(desc).strip() if desc else ""

            up = desc_str.upper()

            # Recurring running-total / carry rows are NOT cash movements.
            # The GL uses several labels across years:
            #   "Balance Forward", "Opening Balance for Year",
            #   "Fiscal Year End", "Account Totals", "5003 Account Totals"
            # These must all be skipped to avoid double-counting the
            # cumulative capital balance each fiscal year.
            _CARRY_MARKERS = (
                "BALANCE FORWARD",
                "OPENING BALANCE FOR YEAR",
                "FISCAL YEAR",
                "ACCOUNT TOTALS",
            )
            if any(mk in up for mk in _CARRY_MARKERS):
                continue

            # The genuine 2017 inception contribution appears once as the
            # "2018 Beginning Balance" / "2017 Imported Beginning" credit.
            is_inception = ("BEGINNING BALAN" in up) or ("IMPORTED BEGINN" in up)

            txn_date = _to_date(entry_dt) or _period_to_date(period)
            if txn_date is None:
                continue

            # amount = debit - credit (contribution credit -> negative)
            amount = float(debit) - float(credit)
            txn_type = "contribution" if credit > 0 else "distribution"
            if is_inception:
                txn_type = "contribution"
                # The inception beginning balance is dated to the period
                # (12/17) not the data-entry date (2019 import artifact).
                p_date = _period_to_date(period)
                if p_date is not None:
                    txn_date = p_date

            cited_amount = Cited(
                value=amount,
                citations=[Citation(
                    source_document_id=source_document_id,
                    locator=Locator(
                        sheet=sheet_name,
                        row=r,
                        note=f"{desc_str[:40]} (Ref {ref})" if ref else desc_str[:40],
                    ),
                    verbatim_text=(f"Dr {debit:,.0f} / Cr {credit:,.0f}"),
                )],
            )

            ledger.transactions.append(EquityTransaction(
                transaction_date=Cited(
                    value=txn_date,
                    citations=[Citation(
                        source_document_id=source_document_id,
                        locator=Locator(sheet=sheet_name, row=r),
                        verbatim_text=str(period) if period else str(txn_date),
                    )],
                ),
                investor_class_id=class_id,
                transaction_type=txn_type,
                amount=cited_amount,
                description=desc_str or None,
                period_label=str(period) if period else None,
                mri_account=sheet_name,
                mri_ref_number=str(ref) if ref else None,
            ))

    wb.close()
    ledger.transactions.sort(key=lambda t: t.transaction_date.value)
    logger.info("Loaded %d equity transactions (KA: %d, IDP: %d)",
                len(ledger.transactions),
                len(ledger.for_investor("KA")),
                len(ledger.for_investor("IDP")))
    return ledger


def annual_equity_summary(
    ledger: EquityLedger,
    class_id: str,
) -> dict[int, dict[str, float]]:
    """Roll the ledger up to an annual {year: {contrib, distrib, net}} summary.

    Contributions are reported as positive cash-in (sign flipped back from
    the internal negative convention) for human-readable presentation.
    """
    out: dict[int, dict[str, float]] = {}
    for t in ledger.for_investor(class_id):
        y = t.transaction_date.value.year
        slot = out.setdefault(y, {"contributions": 0.0,
                                  "distributions": 0.0, "net": 0.0})
        amt = t.amount.value
        if amt < 0:  # contribution (negative internally)
            slot["contributions"] += -amt
        else:        # distribution (positive internally)
            slot["distributions"] += amt
        slot["net"] = slot["contributions"] - slot["distributions"]
    return dict(sorted(out.items()))


__all__ = ["load_equity_ledger", "annual_equity_summary"]
