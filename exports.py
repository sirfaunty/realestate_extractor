"""
CSV and Excel export for Capactive Document Extractor.

Generates downloadable files from extracted data:
- Rent rolls (per-property or all)
- Operating statements
- General ledger entries
- Financial terms
- Clauses
- Property summaries
- Full document index

Uses Python's built-in csv module for CSV and openpyxl for Excel.
Falls back to CSV-only if openpyxl is not installed.
"""

import csv
import io
from datetime import datetime
from typing import List, Dict, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ─── Column Definitions ─────────────────────────────────────────────
#
# Each export type defines its columns as (header, dict_key, format).
# Format: 'text', 'number', 'currency', 'date', 'percent'

RENT_ROLL_COLUMNS = [
    ('Property', 'property_name', 'text'),
    ('Unit', 'unit_number', 'text'),
    ('Suite', 'suite', 'text'),
    ('Tenant', 'tenant_name', 'text'),
    ('Square Footage', 'square_footage', 'number'),
    ('Monthly Rent', 'monthly_rent', 'currency'),
    ('Annual Rent', 'annual_rent', 'currency'),
    ('Rent PSF', 'rent_psf', 'currency'),
    ('Lease Start', 'lease_start', 'date'),
    ('Lease End', 'lease_end', 'date'),
    ('Status', 'status', 'text'),
    ('Notes', 'notes', 'text'),
    ('Source Document', 'filename', 'text'),
]

OPERATING_STATEMENT_COLUMNS = [
    ('Property', 'property_name', 'text'),
    ('Period', 'period', 'text'),
    ('Category', 'category', 'text'),
    ('Subcategory', 'subcategory', 'text'),
    ('Line Item', 'line_item', 'text'),
    ('Amount', 'amount', 'currency'),
    ('Amount PSF', 'amount_psf', 'currency'),
    ('Is Subtotal', 'is_subtotal', 'text'),
    ('Is Total', 'is_total', 'text'),
    ('Source Document', 'filename', 'text'),
]

GL_COLUMNS = [
    ('Property', 'property_name', 'text'),
    ('Date', 'entry_date', 'date'),
    ('Account Code', 'account_code', 'text'),
    ('Account Name', 'account_name', 'text'),
    ('Description', 'description', 'text'),
    ('Debit', 'debit', 'currency'),
    ('Credit', 'credit', 'currency'),
    ('Balance', 'balance', 'currency'),
    ('Period', 'period', 'text'),
    ('Vendor', 'vendor', 'text'),
    ('Reference', 'reference', 'text'),
    ('Source Document', 'filename', 'text'),
]

FINANCIAL_TERMS_COLUMNS = [
    ('Term Type', 'term_type', 'text'),
    ('Label', 'term_label', 'text'),
    ('Value (Raw)', 'value_raw', 'text'),
    ('Value (Numeric)', 'value_numeric', 'number'),
    ('Unit', 'value_unit', 'text'),
    ('Effective Date', 'effective_date', 'date'),
    ('Expiration Date', 'expiration_date', 'date'),
    ('Escalation Type', 'escalation_type', 'text'),
    ('Escalation Detail', 'escalation_detail', 'text'),
    ('Section', 'section_ref', 'text'),
    ('Source Document', 'filename', 'text'),
]

CLAUSES_COLUMNS = [
    ('Clause Type', 'clause_type', 'text'),
    ('Title', 'clause_title', 'text'),
    ('Section', 'section_ref', 'text'),
    ('Summary', 'summary', 'text'),
    ('Full Text', 'full_text', 'text'),
    ('Confidence', 'confidence', 'percent'),
    ('Source Document', 'filename', 'text'),
]

PROPERTY_SUMMARY_COLUMNS = [
    ('Name', 'name', 'text'),
    ('Address', 'address', 'text'),
    ('City', 'city', 'text'),
    ('State', 'state', 'text'),
    ('ZIP', 'zip_code', 'text'),
    ('Type', 'property_type', 'text'),
    ('Status', 'status', 'text'),
    ('Year Built', 'year_built', 'number'),
    ('Total Units', 'total_units', 'number'),
    ('Total SF', 'total_sqft', 'number'),
    ('Acquisition Price', 'acquisition_price', 'currency'),
]

DOCUMENT_INDEX_COLUMNS = [
    ('ID', 'id', 'number'),
    ('Filename', 'filename', 'text'),
    ('Document Type', 'document_type', 'text'),
    ('Property', 'property_name', 'text'),
    ('Pages', 'page_count', 'number'),
    ('Scanned', 'is_scanned', 'text'),
    ('OCR Confidence', 'ocr_confidence', 'percent'),
    ('Review Status', 'review_status', 'text'),
    ('Processed', 'processed_at', 'date'),
]


# ─── CSV Export ──────────────────────────────────────────────────────

def export_csv(rows: List[Dict], columns: list) -> str:
    """
    Export rows to CSV string.

    Args:
        rows: list of dicts with data
        columns: list of (header, key, format) tuples

    Returns CSV as a string.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Header row
    writer.writerow([col[0] for col in columns])

    # Data rows
    for row in rows:
        csv_row = []
        for header, key, fmt in columns:
            val = row.get(key, '')
            if val is None:
                val = ''
            elif fmt == 'currency' and isinstance(val, (int, float)):
                val = round(val, 2)
            elif fmt == 'percent' and isinstance(val, (int, float)):
                val = round(val * 100, 1) if val <= 1 else round(val, 1)
            csv_row.append(val)
        writer.writerow(csv_row)

    return output.getvalue()


def export_csv_bytes(rows: List[Dict], columns: list) -> bytes:
    """Export rows to CSV bytes (UTF-8 with BOM for Excel compatibility)."""
    csv_str = export_csv(rows, columns)
    return b'\xef\xbb\xbf' + csv_str.encode('utf-8')


# ─── Excel Export ────────────────────────────────────────────────────

def export_excel(rows: List[Dict], columns: list,
                 sheet_name: str = 'Export',
                 title: str = None) -> bytes:
    """
    Export rows to Excel (.xlsx) bytes.

    Falls back to CSV if openpyxl is not installed.
    """
    if not HAS_OPENPYXL:
        return export_csv_bytes(rows, columns)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    header_font = Font(name='Inter', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='0B3D6B', end_color='0B3D6B', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='Inter', size=10)
    currency_fmt = '#,##0.00'
    number_fmt = '#,##0'
    percent_fmt = '0.0%'
    thin_border = Border(
        bottom=Side(style='thin', color='E2E4E8'),
    )

    # Title row
    start_row = 1
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name='Inter', bold=True, size=14, color='0B3D6B')
        title_cell.alignment = Alignment(vertical='center')
        ws.row_dimensions[1].height = 30

        # Subtitle with export date
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
        sub_cell = ws.cell(row=2, column=1,
                          value=f'Exported {datetime.now().strftime("%B %d, %Y at %I:%M %p")}')
        sub_cell.font = Font(name='Inter', size=9, color='8E8E9A')
        start_row = 4

    # Header row
    for col_idx, (header, key, fmt) in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    ws.row_dimensions[start_row].height = 28

    # Data rows
    for row_idx, row in enumerate(rows, start_row + 1):
        for col_idx, (header, key, fmt) in enumerate(columns, 1):
            val = row.get(key, '')
            if val is None:
                val = ''

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border

            # Apply format
            if fmt == 'currency' and isinstance(val, (int, float)):
                cell.number_format = currency_fmt
                cell.alignment = Alignment(horizontal='right')
            elif fmt == 'number' and isinstance(val, (int, float)):
                cell.number_format = number_fmt
                cell.alignment = Alignment(horizontal='right')
            elif fmt == 'percent' and isinstance(val, (int, float)):
                cell.number_format = percent_fmt
                cell.alignment = Alignment(horizontal='right')

        # Alternate row shading
        if row_idx % 2 == 0:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color='F7F8FA', end_color='F7F8FA', fill_type='solid')

    # Auto-width columns (cap at 40)
    for col_idx, (header, key, fmt) in enumerate(columns, 1):
        max_len = len(header)
        for row_idx in range(start_row + 1, min(start_row + 51, start_row + len(rows) + 1)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, min(len(str(cell_val)), 40))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 3

    # Freeze header row
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # Auto-filter
    ws.auto_filter.ref = f'A{start_row}:{openpyxl.utils.get_column_letter(len(columns))}{start_row + len(rows)}'

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ─── Multi-Sheet Excel Export ────────────────────────────────────────

def export_property_workbook(property_name: str, data: Dict) -> bytes:
    """
    Export a full property workbook with multiple sheets:
    - Summary
    - Rent Roll
    - Operating Statement
    - Financial Terms
    - GL Entries
    - Clauses
    """
    if not HAS_OPENPYXL:
        # Fallback: just export rent roll as CSV
        return export_csv_bytes(data.get('rent_roll', []), RENT_ROLL_COLUMNS)

    wb = openpyxl.Workbook()

    sheets = [
        ('Rent Roll', data.get('rent_roll', []), RENT_ROLL_COLUMNS),
        ('Operating Statement', data.get('operating_statement', []), OPERATING_STATEMENT_COLUMNS),
        ('Financial Terms', data.get('financial_terms', []), FINANCIAL_TERMS_COLUMNS),
        ('GL Entries', data.get('gl_entries', []), GL_COLUMNS),
        ('Clauses', data.get('clauses', []), CLAUSES_COLUMNS),
    ]

    first = True
    for sheet_name, rows, columns in sheets:
        if not rows:
            continue
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        _write_sheet(ws, rows, columns, title=f'{property_name} — {sheet_name}')

    # If no data at all, write an empty summary sheet
    if first:
        ws = wb.active
        ws.title = 'Summary'
        ws.cell(row=1, column=1, value=f'{property_name} — No data to export')
        ws.cell(row=1, column=1).font = Font(name='Inter', bold=True, size=14, color='0B3D6B')

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _write_sheet(ws, rows, columns, title=None):
    """Write data to a worksheet with Capactive styling."""
    header_font = Font(name='Inter', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='0B3D6B', end_color='0B3D6B', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='Inter', size=10)
    currency_fmt = '#,##0.00'
    number_fmt = '#,##0'
    percent_fmt = '0.0%'
    thin_border = Border(bottom=Side(style='thin', color='E2E4E8'))

    start_row = 1
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        tc = ws.cell(row=1, column=1, value=title)
        tc.font = Font(name='Inter', bold=True, size=14, color='0B3D6B')
        ws.row_dimensions[1].height = 30

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
        sc = ws.cell(row=2, column=1,
                    value=f'Exported {datetime.now().strftime("%B %d, %Y")} | {len(rows)} rows')
        sc.font = Font(name='Inter', size=9, color='8E8E9A')
        start_row = 4

    for col_idx, (header, key, fmt) in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    ws.row_dimensions[start_row].height = 28

    for row_idx, row in enumerate(rows, start_row + 1):
        for col_idx, (header, key, fmt) in enumerate(columns, 1):
            val = row.get(key, '')
            if val is None:
                val = ''
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if fmt == 'currency' and isinstance(val, (int, float)):
                cell.number_format = currency_fmt
                cell.alignment = Alignment(horizontal='right')
            elif fmt == 'number' and isinstance(val, (int, float)):
                cell.number_format = number_fmt
                cell.alignment = Alignment(horizontal='right')
            elif fmt == 'percent' and isinstance(val, (int, float)):
                cell.number_format = percent_fmt
                cell.alignment = Alignment(horizontal='right')

        if row_idx % 2 == 0:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color='F7F8FA', end_color='F7F8FA', fill_type='solid')

    for col_idx, (header, key, fmt) in enumerate(columns, 1):
        max_len = len(header)
        for ri in range(start_row + 1, min(start_row + 51, start_row + len(rows) + 1)):
            cv = ws.cell(row=ri, column=col_idx).value
            if cv:
                max_len = max(max_len, min(len(str(cv)), 40))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 3

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    ws.auto_filter.ref = f'A{start_row}:{openpyxl.utils.get_column_letter(len(columns))}{start_row + len(rows)}'


# ─── Export Type Registry ────────────────────────────────────────────

EXPORT_TYPES = {
    'rent_roll': {
        'label': 'Rent Roll',
        'columns': RENT_ROLL_COLUMNS,
        'db_method': 'get_rent_roll',
    },
    'operating_statement': {
        'label': 'Operating Statement',
        'columns': OPERATING_STATEMENT_COLUMNS,
        'db_method': 'get_operating_statement',
    },
    'gl_entries': {
        'label': 'General Ledger',
        'columns': GL_COLUMNS,
        'db_method': 'get_gl_entries',
    },
    'financial_terms': {
        'label': 'Financial Terms',
        'columns': FINANCIAL_TERMS_COLUMNS,
        'db_method': 'get_financial_terms',
    },
    'clauses': {
        'label': 'Clauses',
        'columns': CLAUSES_COLUMNS,
        'db_method': 'get_clauses',
    },
    'properties': {
        'label': 'Properties',
        'columns': PROPERTY_SUMMARY_COLUMNS,
        'db_method': 'list_properties',
    },
    'documents': {
        'label': 'Document Index',
        'columns': DOCUMENT_INDEX_COLUMNS,
        'db_method': 'list_documents',
    },
}
