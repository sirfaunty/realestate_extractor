"""
Monthly Report Package Parser for Village Green / property management exports.

Handles multi-sheet Excel workbooks from .msg email packages containing:
  - Executive Summary Report (unit mix, renewals, NTV, delinquency, occupancy)
  - 12-Month Statement (full P&L with monthly columns)
  - Variance Report (budget vs actual with PTD/YTD)

These reports are typically emailed as .msg attachments and contain
structured Excel data that maps directly to rent_roll_entries and
operating_statement_items tables.
"""

import json
import logging
import os
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════
#  MSG Extraction
# ═══════════════════════════════════════════════════════════════════════

def extract_msg_attachments(msg_path: str, dest_dir: str) -> List[Dict[str, str]]:
    """
    Extract file attachments from an Outlook .msg file.

    Returns list of dicts: {'filename': ..., 'path': ..., 'size': ...}
    """
    try:
        import extract_msg
    except ImportError:
        logger.warning("extract-msg not installed — run: pip install extract-msg")
        return []

    attachments = []
    try:
        msg = extract_msg.Message(msg_path)
        for att in msg.attachments:
            if not att.longFilename:
                continue
            fname = att.longFilename
            out_path = os.path.join(dest_dir, fname)
            with open(out_path, 'wb') as f:
                f.write(att.data)
            attachments.append({
                'filename': fname,
                'path': out_path,
                'size': len(att.data),
            })
        msg.close()
    except Exception as e:
        logger.error(f"Failed to extract .msg attachments: {e}")

    return attachments


# ═══════════════════════════════════════════════════════════════════════
#  Utility helpers
# ═══════════════════════════════════════════════════════════════════════

def _safe_float(val) -> Optional[float]:
    """Convert a value to float, handling None, strings, percentages."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val = val.strip().replace(',', '').replace('$', '').replace('%', '')
        if val.startswith('(') and val.endswith(')'):
            val = '-' + val[1:-1]
        try:
            return float(val)
        except (ValueError, TypeError):
            return None
    return None


def _parse_period_from_filename(filename: str) -> Optional[str]:
    """
    Extract period from filename like 'CBL 03.2026 ...' → 'Mar 2026'.
    """
    m = re.search(r'(\d{2})\.(\d{4})', filename)
    if m:
        month_num = int(m.group(1))
        year = m.group(2)
        months = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                  'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        if 1 <= month_num <= 12:
            return f"{months[month_num]} {year}"
    return None


def _detect_property_name(ws, max_rows: int = 10) -> Optional[str]:
    """Try to find property name in the first few rows of a worksheet."""
    # Strategy 1: Look for "Property:" label followed by name (Executive Summary format)
    for row in ws.iter_rows(min_row=1, max_row=max_rows, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value.strip().lower()
                if val == 'property:' or val == 'property':
                    # Name is in the next column
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1).value
                    if next_cell and isinstance(next_cell, str) and next_cell.strip():
                        return next_cell.strip()

    # Strategy 2: Look for "The <Name> (<code>)" pattern in row 1
    row1 = ws.cell(row=1, column=1).value
    if row1 and isinstance(row1, str):
        m = re.match(r'^(The\s+\w[\w\s]+?)(?:\s*\(.*\))?\s*$', row1.strip())
        if m:
            return m.group(1).strip()

    # Strategy 3: First non-header text in rows 1-5
    _skip = {'statement', 'budget', 'report', 'period', 'book', 'village green',
             'comparison', 'notes', 'formulas', 'beginning'}
    for row in ws.iter_rows(min_row=1, max_row=max_rows, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str):
                val = cell.strip()
                if val and len(val) > 3 and not any(s in val.lower() for s in _skip):
                    return val
    return None


# ═══════════════════════════════════════════════════════════════════════
#  Executive Summary Report — Unit Mix
# ═══════════════════════════════════════════════════════════════════════

def parse_unit_mix(ws) -> List[Dict[str, Any]]:
    """
    Parse 'Exhibit 6 Unit Mix' sheet into rent_roll_entries.

    Expected layout (Village Green format):
      Row 9:  property name, date
      Row 10: Header row (Style, Size, Quantity, Average Rent, ...)
      Row 11+: Data rows
    """
    entries = []

    # Find the header row by looking for 'Style' in column B
    header_row = None
    for row_idx in range(1, 20):
        cell = ws.cell(row=row_idx, column=2).value
        if cell and isinstance(cell, str) and 'style' in cell.lower():
            header_row = row_idx
            break

    if not header_row:
        logger.warning("Unit Mix: could not find header row")
        return entries

    # Map column indices from the header
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val and isinstance(val, str):
            headers[col_idx] = val.strip().lower()

    # Read data rows until we hit an empty Style cell or 'Total'
    for row_idx in range(header_row + 1, ws.max_row + 1):
        style = ws.cell(row=row_idx, column=2).value
        if not style:
            continue
        style_str = str(style).strip()
        if not style_str or style_str.lower() in ('total', 'totals', 'grand total'):
            # Check if it's a total row — capture it but mark it
            qty_val = ws.cell(row=row_idx, column=4).value
            if style_str.lower().startswith('total') and qty_val is not None:
                # Skip total rows for rent_roll but they might be useful as metadata
                continue
            if not style_str:
                continue

        # Read the row values by column position
        size = ws.cell(row=row_idx, column=3).value  # bed/bath
        quantity = _safe_float(ws.cell(row=row_idx, column=4).value)
        avg_rent = _safe_float(ws.cell(row=row_idx, column=5).value)
        occupied = _safe_float(ws.cell(row=row_idx, column=6).value)
        vacant = _safe_float(ws.cell(row=row_idx, column=7).value)
        ntv = _safe_float(ws.cell(row=row_idx, column=8).value)
        pct_occ = _safe_float(ws.cell(row=row_idx, column=9).value)
        pct_vacant = _safe_float(ws.cell(row=row_idx, column=10).value)
        pct_avail = _safe_float(ws.cell(row=row_idx, column=11).value)

        if quantity is None and avg_rent is None:
            continue

        # Parse bed/bath from size string like "1 bed / 1 bath"
        beds, baths = None, None
        if size and isinstance(size, str):
            m = re.match(r'(\d+)\s*bed\s*/\s*(\d+)\s*bath', size, re.IGNORECASE)
            if m:
                beds, baths = int(m.group(1)), int(m.group(2))

        # Build a rent_roll_entry per unit style.
        # Since this is a unit mix summary (aggregated), we store one entry per style
        # with quantity in metadata.
        entry = {
            'unit_number': style_str,
            'tenant_name': None,  # unit mix is aggregate, no individual tenants
            'monthly_rent': avg_rent,
            'annual_rent': avg_rent * 12 if avg_rent else None,
            'square_footage': None,  # not in unit mix
            'status': _unit_mix_status(occupied, vacant, int(quantity or 0)),
            'notes': f"{size or ''} | {int(quantity or 0)} units",
            'metadata': json.dumps({
                'source': 'unit_mix',
                'style': style_str,
                'size': str(size) if size else None,
                'beds': beds,
                'baths': baths,
                'quantity': int(quantity) if quantity else None,
                'occupied': int(occupied) if occupied else None,
                'vacant': int(vacant) if vacant else None,
                'ntv': int(ntv) if ntv else None,
                'pct_occupied': round(pct_occ * 100, 1) if pct_occ else None,
                'pct_vacant': round(pct_vacant * 100, 1) if pct_vacant else None,
                'pct_available': round(pct_avail * 100, 1) if pct_avail else None,
            }),
        }
        entries.append(entry)

    logger.info(f"Unit Mix: parsed {len(entries)} unit styles")
    return entries


def _unit_mix_status(occupied, vacant, total) -> str:
    """Derive overall status string from counts."""
    if total and occupied:
        ratio = occupied / total
        if ratio >= 0.95:
            return 'occupied'
        elif ratio >= 0.5:
            return 'partially_occupied'
        else:
            return 'mostly_vacant'
    return 'unknown'


# ═══════════════════════════════════════════════════════════════════════
#  Executive Summary Report — Renewals
# ═══════════════════════════════════════════════════════════════════════

def parse_renewals(ws) -> List[Dict[str, Any]]:
    """
    Parse 'Exhibit 4 Renewals' sheet.

    Returns list of dicts with per-unit renewal info.
    """
    entries = []

    # Find header row (Resident, Unit #, ...)
    header_row = None
    for row_idx in range(1, 15):
        cell = ws.cell(row=row_idx, column=2).value
        if cell and isinstance(cell, str) and 'resident' in cell.lower():
            header_row = row_idx
            break

    if not header_row:
        logger.warning("Renewals: could not find header row")
        return entries

    for row_idx in range(header_row + 1, ws.max_row + 1):
        resident = ws.cell(row=row_idx, column=2).value
        if not resident or not isinstance(resident, str) or not resident.strip():
            continue
        if resident.strip().lower() in ('total', 'totals', 'grand total', 'averages'):
            continue

        unit_num = ws.cell(row=row_idx, column=3).value
        bed_bath = ws.cell(row=row_idx, column=4).value
        unit_style = ws.cell(row=row_idx, column=5).value
        sqft = _safe_float(ws.cell(row=row_idx, column=6).value)
        current_rent = _safe_float(ws.cell(row=row_idx, column=7).value)
        current_psf = _safe_float(ws.cell(row=row_idx, column=8).value)
        renewed_rent = _safe_float(ws.cell(row=row_idx, column=9).value)
        renewed_psf = _safe_float(ws.cell(row=row_idx, column=10).value)
        delta_pct = _safe_float(ws.cell(row=row_idx, column=11).value)
        delta_total = _safe_float(ws.cell(row=row_idx, column=12).value)
        decision = ws.cell(row=row_idx, column=13).value

        entry = {
            'unit_number': str(unit_num).strip() if unit_num else None,
            'tenant_name': resident.strip(),
            'monthly_rent': renewed_rent or current_rent,
            'annual_rent': (renewed_rent or current_rent or 0) * 12 if (renewed_rent or current_rent) else None,
            'square_footage': sqft,
            'rent_psf': renewed_psf or current_psf,
            'status': _renewal_status(decision),
            'notes': str(decision).strip() if decision else None,
            'metadata': json.dumps({
                'source': 'renewals',
                'bed_bath': str(bed_bath) if bed_bath else None,
                'unit_style': str(unit_style) if unit_style else None,
                'current_rent': current_rent,
                'renewed_rent': renewed_rent,
                'delta_pct': round(delta_pct * 100, 1) if delta_pct else None,
                'delta_total': delta_total,
                'decision': str(decision).strip() if decision else None,
            }),
        }
        entries.append(entry)

    logger.info(f"Renewals: parsed {len(entries)} entries")
    return entries


def _renewal_status(decision) -> str:
    if not decision:
        return 'unknown'
    d = str(decision).lower().strip()
    if 'renew' in d:
        return 'occupied'
    if 'move' in d or 'transfer' in d or 'community' in d:
        return 'move_out'
    return 'unknown'


# ═══════════════════════════════════════════════════════════════════════
#  Executive Summary Report — NTV (Notice to Vacate)
# ═══════════════════════════════════════════════════════════════════════

def parse_ntv(ws) -> List[Dict[str, Any]]:
    """
    Parse 'Exhibit 3 NTV' sheet.

    Returns list of dicts with per-unit notice-to-vacate info.
    """
    entries = []

    header_row = None
    for row_idx in range(1, 15):
        cell = ws.cell(row=row_idx, column=2).value
        if cell and isinstance(cell, str) and 'resident' in cell.lower():
            header_row = row_idx
            break

    if not header_row:
        logger.warning("NTV: could not find header row")
        return entries

    for row_idx in range(header_row + 1, ws.max_row + 1):
        resident = ws.cell(row=row_idx, column=2).value
        if not resident or not isinstance(resident, str) or not resident.strip():
            continue
        if resident.strip().lower() in ('total', 'totals', 'grand total'):
            continue

        unit_num = ws.cell(row=row_idx, column=3).value
        bed_bath = ws.cell(row=row_idx, column=4).value
        unit_style = ws.cell(row=row_idx, column=5).value
        sqft = _safe_float(ws.cell(row=row_idx, column=6).value)
        current_rent = _safe_float(ws.cell(row=row_idx, column=7).value)
        proposed_rent = _safe_float(ws.cell(row=row_idx, column=8).value)
        delta_pct = _safe_float(ws.cell(row=row_idx, column=9).value)
        delta_total = _safe_float(ws.cell(row=row_idx, column=10).value)
        mo_date = ws.cell(row=row_idx, column=11).value
        reason = ws.cell(row=row_idx, column=12).value

        entry = {
            'unit_number': str(unit_num).strip() if unit_num else None,
            'tenant_name': resident.strip(),
            'monthly_rent': current_rent,
            'annual_rent': current_rent * 12 if current_rent else None,
            'square_footage': sqft,
            'status': 'notice_to_vacate',
            'lease_end': mo_date.strftime('%Y-%m-%d') if isinstance(mo_date, datetime) else str(mo_date) if mo_date else None,
            'notes': f"NTV: {reason}" if reason else "Notice to Vacate",
            'metadata': json.dumps({
                'source': 'ntv',
                'bed_bath': str(bed_bath) if bed_bath else None,
                'unit_style': str(unit_style) if unit_style else None,
                'current_rent': current_rent,
                'proposed_rent': proposed_rent,
                'delta_pct': round(delta_pct * 100, 1) if delta_pct else None,
                'move_out_date': mo_date.strftime('%Y-%m-%d') if isinstance(mo_date, datetime) else str(mo_date) if mo_date else None,
                'reason': str(reason).strip() if reason else None,
            }),
        }
        entries.append(entry)

    logger.info(f"NTV: parsed {len(entries)} entries")
    return entries


# ═══════════════════════════════════════════════════════════════════════
#  Executive Summary Report — Delinquency
# ═══════════════════════════════════════════════════════════════════════

def parse_delinquency(ws) -> List[Dict[str, Any]]:
    """
    Parse 'Exhibit 5 Delinquency' sheet → financial terms.

    Returns list of dicts suitable for financial_terms table.
    """
    terms = []

    header_row = None
    for row_idx in range(1, 15):
        cell = ws.cell(row=row_idx, column=2).value
        if cell and isinstance(cell, str) and 'month' in cell.lower():
            header_row = row_idx
            break

    if not header_row:
        logger.warning("Delinquency: could not find header row")
        return terms

    for row_idx in range(header_row + 1, ws.max_row + 1):
        month = ws.cell(row=row_idx, column=2).value
        if not month:
            continue
        month_str = month.strftime('%b %Y') if isinstance(month, datetime) else str(month).strip()
        if not month_str or month_str.lower() in ('total', 'average'):
            continue

        num_delinquent = _safe_float(ws.cell(row=row_idx, column=3).value)
        total_dq = _safe_float(ws.cell(row=row_idx, column=4).value)
        past_due_60 = _safe_float(ws.cell(row=row_idx, column=5).value)
        gpr = _safe_float(ws.cell(row=row_idx, column=6).value)
        pct_dq = _safe_float(ws.cell(row=row_idx, column=7).value)

        if total_dq is not None:
            terms.append({
                'term_type': 'delinquency_total',
                'term_label': f'Total Delinquency ({month_str})',
                'value_raw': f'${total_dq:,.2f}' if total_dq else '0',
                'value_numeric': total_dq,
                'period': month_str,
                'metadata': json.dumps({
                    'source': 'delinquency',
                    'num_delinquent': int(num_delinquent) if num_delinquent else None,
                    'past_due_60_plus': past_due_60,
                    'gpr': gpr,
                    'pct_dq_to_gpr': round(pct_dq * 100, 2) if pct_dq else None,
                }),
            })

    logger.info(f"Delinquency: parsed {len(terms)} monthly entries")
    return terms


# ═══════════════════════════════════════════════════════════════════════
#  Executive Summary Report — Exhibit 1 Summary (Occupancy)
# ═══════════════════════════════════════════════════════════════════════

def parse_occupancy_summary(ws) -> List[Dict[str, Any]]:
    """
    Parse 'Exhibit 1 Summary' sheet for occupancy and GPR metrics.

    Village Green format:
      Row 6: "Physical Occupancy" header
      Row 7: Current / 30-Day / 60-Day labels
      Row 8: Values (decimals like 0.768)
      Row 16: "12 Month Trailing" header with Market Rent / GPR / NRI columns
      Row 17-28: Monthly rows

    Returns list of financial term dicts.
    """
    terms = []

    # ── Physical Occupancy (rows 6-8 area) ──
    for row_idx in range(1, 15):
        cell = ws.cell(row=row_idx, column=2).value
        if not cell or not isinstance(cell, str):
            continue
        if 'physical occupancy' in cell.lower():
            # Next row has labels, row after has values
            label_row = row_idx + 1
            value_row = row_idx + 2
            for col in range(2, 8):
                label = ws.cell(row=label_row, column=col).value
                val = ws.cell(row=value_row, column=col).value
                num = _safe_float(val)
                if label and num is not None:
                    # Convert decimal to percentage
                    if num <= 1:
                        num = round(num * 100, 1)
                    label_str = str(label).strip()
                    terms.append({
                        'term_type': 'occupancy_rate',
                        'term_label': f'Physical Occupancy ({label_str})',
                        'value_raw': f'{num}%',
                        'value_numeric': num,
                    })
            break

    # ── 12-Month Trailing table ──
    for row_idx in range(1, 25):
        cell = ws.cell(row=row_idx, column=2).value
        if not cell or not isinstance(cell, str):
            continue
        if '12 month trailing' in cell.lower():
            # Header row has column labels
            header_row = row_idx
            # Map the column labels (Market Rent, GPR, NRI)
            col_labels = {}
            for col in range(3, 12):
                hdr = ws.cell(row=header_row, column=col).value
                if hdr and isinstance(hdr, str):
                    col_labels[col] = hdr.strip()

            # Read monthly rows
            for data_row in range(header_row + 1, header_row + 14):
                month_label = ws.cell(row=data_row, column=2).value
                if not month_label:
                    continue
                month_str = str(month_label).strip()
                if not month_str or month_str.lower() in ('total', 'average'):
                    continue

                for col, col_label in col_labels.items():
                    val = _safe_float(ws.cell(row=data_row, column=col).value)
                    if val is None:
                        continue

                    label_lower = col_label.lower()
                    if 'market rent' in label_lower:
                        term_type = 'market_rent'
                    elif 'gross potential' in label_lower or 'gpr' in label_lower:
                        term_type = 'gross_potential_rent'
                    elif 'net rental' in label_lower or 'nri' in label_lower:
                        term_type = 'net_rental_income'
                    else:
                        term_type = 'financial_metric'

                    terms.append({
                        'term_type': term_type,
                        'term_label': f'{col_label} ({month_str})',
                        'value_raw': f'${val:,.0f}',
                        'value_numeric': val,
                        'period': month_str,
                        'metadata': json.dumps({'source': 'occupancy_summary_trailing'}),
                    })
            break

    # ── Scan for any additional labeled metrics ──
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, min(ws.max_column + 1, 20)):
            cell = ws.cell(row=row_idx, column=col_idx).value
            if not cell or not isinstance(cell, str):
                continue
            label = cell.strip().lower()

            # Catch metrics not handled above
            if label == 'gpr' or label == 'gpr:':
                val = ws.cell(row=row_idx, column=col_idx + 1).value
                num = _safe_float(val)
                if num is not None and num > 1000:  # looks like a dollar amount
                    terms.append({
                        'term_type': 'gross_potential_rent',
                        'term_label': 'GPR (Executive Summary)',
                        'value_raw': f'${num:,.0f}',
                        'value_numeric': num,
                    })

    logger.info(f"Occupancy Summary: parsed {len(terms)} metrics")
    return terms


# ═══════════════════════════════════════════════════════════════════════
#  Executive Summary sheet — top-level property metrics
# ═══════════════════════════════════════════════════════════════════════

def _parse_exec_summary_metrics(ws) -> List[Dict[str, Any]]:
    """
    Parse the Executive Summary sheet for property-level metrics.

    VG format row 3: Property: <name> ... Units: <n> ... GPR: <amount>
    """
    terms = []

    # Scan first 10 rows for labeled metrics
    _METRIC_MAP = {
        'units:': ('total_units', 'Total Units'),
        'units': ('total_units', 'Total Units'),
        'gpr:': ('gross_potential_rent', 'GPR (Monthly)'),
        'gpr': ('gross_potential_rent', 'GPR (Monthly)'),
        'year built:': ('year_built', 'Year Built'),
        'year built': ('year_built', 'Year Built'),
    }

    for row_idx in range(1, 12):
        for col_idx in range(1, 20):
            cell = ws.cell(row=row_idx, column=col_idx).value
            if not cell or not isinstance(cell, str):
                continue
            key = cell.strip().lower()
            if key in _METRIC_MAP:
                term_type, label = _METRIC_MAP[key]
                val = ws.cell(row=row_idx, column=col_idx + 1).value
                num = _safe_float(val)
                if num is not None:
                    terms.append({
                        'term_type': term_type,
                        'term_label': label,
                        'value_raw': f'{int(num)}' if term_type in ('total_units', 'year_built') else f'${num:,.0f}',
                        'value_numeric': num,
                        'metadata': json.dumps({'source': 'executive_summary'}),
                    })

    logger.info(f"Executive Summary metrics: parsed {len(terms)} terms")
    return terms


# ═══════════════════════════════════════════════════════════════════════
#  12-Month Statement → operating_statement_items
# ═══════════════════════════════════════════════════════════════════════

# Category classification based on section headers in VG P&L format
_OS_CATEGORY_MAP = {
    'rental income': 'revenue',
    'income adjustments': 'revenue',
    'total rental income': 'revenue',
    'other income': 'other_income',
    'utility income': 'utility_income',
    'total income': 'revenue',
    'variable expenses': 'operating_expense',
    'payroll': 'operating_expense',
    'management fees': 'management_fee',
    'administrative': 'operating_expense',
    'grounds': 'operating_expense',
    'marketing': 'operating_expense',
    'maintenance': 'operating_expense',
    'security': 'operating_expense',
    'apartment turnover': 'operating_expense',
    'total variable expenses': 'operating_expense',
    'fixed expenses': 'fixed_expense',
    'utilities': 'utility_expense',
    'taxes': 'tax',
    'insurance': 'insurance',
    'total fixed expenses': 'fixed_expense',
    'total operating expenses': 'operating_expense',
    'net operating income': 'noi',
    'major expense': 'capital',
    'noi after major expense': 'noi',
    'debt service': 'debt_service',
    'noi after debt services': 'noi',
    'depreciation & amortization': 'depreciation',
    'depreciation / amortization': 'depreciation',
    'other expenses': 'other_expense',
    'net income (loss)': 'net_income',
    'capital improvements': 'capital',
    'escrows': 'escrow',
    'owners contribution / distribution': 'distribution',
    'other changes in balance sheet': 'balance_sheet',
    'cash flow': 'cash_flow',
}


def parse_12month_statement(ws, property_name: str = None) -> List[Dict[str, Any]]:
    """
    Parse a Village Green 12-Month Statement into operating_statement_items.

    Layout:
      Row 1: Property name
      Row 2: "Statement (12 months)"
      Row 3: "Period = Apr 2025-Mar 2026"
      Row 5: Month headers (col C onwards)
      Row 6+: Category headers (col B only) and line items (col A = acct code, col B = name)
    """
    items = []

    # Get property name from sheet if not provided
    if not property_name:
        property_name = _detect_property_name(ws)

    # Read month headers from row 5
    month_headers = {}  # col_idx → period string
    for col_idx in range(3, ws.max_column + 1):
        val = ws.cell(row=5, column=col_idx).value
        if val:
            month_headers[col_idx] = str(val).strip()

    if not month_headers:
        logger.warning("12-Month Statement: no month headers found in row 5")
        return items

    logger.info(f"12-Month Statement: found {len(month_headers)} period columns: "
                f"{list(month_headers.values())[:3]}...{list(month_headers.values())[-1:]}")

    # Track current category/subcategory
    current_category = 'revenue'
    current_subcategory = None

    for row_idx in range(6, ws.max_row + 1):
        acct_code = ws.cell(row=row_idx, column=1).value
        line_name = ws.cell(row=row_idx, column=2).value

        if not line_name:
            continue

        line_name_str = str(line_name).strip()
        if not line_name_str:
            continue

        line_lower = line_name_str.lower()

        # Is this a category header (no account code, no amounts)?
        is_header = (acct_code is None)
        is_total = line_lower.startswith('total') or line_lower.startswith('net ')
        is_subtotal = is_total and not line_lower.startswith('total income') and not line_lower.startswith('total operating')

        # Update category tracking
        if is_header:
            cat_key = line_lower.replace('total ', '').strip()
            if cat_key in _OS_CATEGORY_MAP:
                if not is_total:
                    current_subcategory = line_name_str
                current_category = _OS_CATEGORY_MAP.get(line_lower, current_category)
            elif line_lower in _OS_CATEGORY_MAP:
                current_category = _OS_CATEGORY_MAP[line_lower]

        # Read amounts for each month column
        for col_idx, period in month_headers.items():
            amount = _safe_float(ws.cell(row=row_idx, column=col_idx).value)
            if amount is None:
                continue

            item = {
                'category': current_category,
                'subcategory': current_subcategory,
                'line_item': line_name_str,
                'amount': amount,
                'period': period,
                'is_subtotal': is_subtotal,
                'is_total': is_total and not is_subtotal,
                'property_name': property_name,
                'metadata': json.dumps({
                    'source': '12month_statement',
                    'account_code': str(acct_code).strip() if acct_code else None,
                }),
            }
            items.append(item)

    logger.info(f"12-Month Statement: parsed {len(items)} items "
                f"across {len(month_headers)} periods")
    return items


# ═══════════════════════════════════════════════════════════════════════
#  Variance Report → operating_statement_items
# ═══════════════════════════════════════════════════════════════════════

def parse_variance_report(ws, property_name: str = None) -> List[Dict[str, Any]]:
    """
    Parse a Village Green Variance Report (Budget Comparison).

    Handles two formats:

    NEW FORMAT (Oct 2025+):
      Row 1: Property name ("The Chamberlain (cbl)")
      Row 2: "Budget Comparison"
      Row 3: "Period = Mar 2026"
      Row 5: Headers (PTD Actual, PTD Budget, Variance, % Var, YTD Actual, ...)
      Row 6+: Account code in col A, line item in col B, amounts col C+

    OLD FORMAT (pre-Oct 2025):
      Row 1: Long title/explanation text
      Row 2: Section headers ("Current Period" / "Year-To-Date")
      Row 3: Headers (Actual, Budget, Variance, %, ...)
      Row 4+: Line item in col A, amounts in col B+
    """
    items = []

    if not property_name:
        property_name = _detect_property_name(ws)

    # Detect format by checking row 5 for header-like text
    format_type = _detect_variance_format(ws)
    logger.info(f"Variance Report: detected {format_type} format")

    if format_type == 'new':
        items = _parse_variance_new_format(ws, property_name)
    else:
        items = _parse_variance_old_format(ws, property_name)

    return items


def _detect_variance_format(ws) -> str:
    """Detect whether this is the new (Oct 2025+) or old variance report format."""
    # New format: row 5 has text headers like "PTD Actual"
    for col in range(3, 15):
        val = ws.cell(row=5, column=col).value
        if val and isinstance(val, str) and any(
            kw in val.lower() for kw in ['ptd', 'ytd', 'actual', 'budget', 'annual']
        ):
            return 'new'

    # Old format: row 3 has headers like "Actual", "Budget", "Variance"
    for col in range(2, 12):
        val = ws.cell(row=3, column=col).value
        if val and isinstance(val, str) and val.strip().lower() in (
            'actual', 'budget', 'variance', '%'
        ):
            return 'old'

    # Default to new format
    return 'new'


def _parse_variance_new_format(ws, property_name: str = None) -> List[Dict[str, Any]]:
    """Parse new-style variance report (Oct 2025+ with account codes)."""
    items = []

    # Parse period from row 3
    period_row = ws.cell(row=3, column=1).value
    report_period = None
    if period_row and isinstance(period_row, str):
        m = re.search(r'Period\s*=\s*(.+)', period_row, re.IGNORECASE)
        if m:
            report_period = m.group(1).strip()

    # Column mapping from row 5
    col_map = {}
    for col_idx in range(3, ws.max_column + 1):
        val = ws.cell(row=5, column=col_idx).value
        if val:
            col_map[col_idx] = str(val).strip()

    logger.info(f"Variance Report: columns = {col_map}")

    current_category = 'revenue'
    current_subcategory = None

    for row_idx in range(6, ws.max_row + 1):
        acct_code = ws.cell(row=row_idx, column=1).value
        line_name = ws.cell(row=row_idx, column=2).value

        if not line_name:
            continue

        line_name_str = str(line_name).strip()
        if not line_name_str:
            continue

        line_lower = line_name_str.lower()
        is_header = (acct_code is None)
        is_total = line_lower.startswith('total') or line_lower.startswith('net ')

        if is_header:
            cat_key = line_lower.replace('total ', '').strip()
            if cat_key in _OS_CATEGORY_MAP:
                if not is_total:
                    current_subcategory = line_name_str
                current_category = _OS_CATEGORY_MAP.get(line_lower, current_category)
            elif line_lower in _OS_CATEGORY_MAP:
                current_category = _OS_CATEGORY_MAP[line_lower]

        ptd_actual = _safe_float(ws.cell(row=row_idx, column=3).value)
        ptd_budget = _safe_float(ws.cell(row=row_idx, column=4).value)
        ptd_variance = _safe_float(ws.cell(row=row_idx, column=5).value)
        ptd_pct_var = _safe_float(ws.cell(row=row_idx, column=6).value)
        ytd_actual = _safe_float(ws.cell(row=row_idx, column=7).value)
        ytd_budget = _safe_float(ws.cell(row=row_idx, column=8).value)
        ytd_variance = _safe_float(ws.cell(row=row_idx, column=9).value)
        ytd_pct_var = _safe_float(ws.cell(row=row_idx, column=10).value)
        annual_budget = _safe_float(ws.cell(row=row_idx, column=11).value)
        note = ws.cell(row=row_idx, column=12).value

        if ptd_actual is None and ytd_actual is None and annual_budget is None:
            continue

        is_subtotal = is_total

        if ptd_actual is not None:
            items.append({
                'category': current_category,
                'subcategory': current_subcategory,
                'line_item': line_name_str,
                'amount': ptd_actual,
                'period': f'{report_period} PTD' if report_period else 'PTD',
                'is_subtotal': is_subtotal,
                'is_total': False,
                'property_name': property_name,
                'metadata': json.dumps({
                    'source': 'variance_report',
                    'account_code': str(acct_code).strip() if acct_code else None,
                    'budget': ptd_budget,
                    'variance': ptd_variance,
                    'pct_variance': ptd_pct_var,
                    'note': str(note).strip() if note else None,
                }),
            })

        if ytd_actual is not None:
            items.append({
                'category': current_category,
                'subcategory': current_subcategory,
                'line_item': line_name_str,
                'amount': ytd_actual,
                'period': f'{report_period} YTD' if report_period else 'YTD',
                'is_subtotal': is_subtotal,
                'is_total': False,
                'property_name': property_name,
                'metadata': json.dumps({
                    'source': 'variance_report',
                    'account_code': str(acct_code).strip() if acct_code else None,
                    'budget': ytd_budget,
                    'variance': ytd_variance,
                    'pct_variance': ytd_pct_var,
                    'note': str(note).strip() if note else None,
                }),
            })

        if annual_budget is not None:
            items.append({
                'category': current_category,
                'subcategory': current_subcategory,
                'line_item': line_name_str,
                'amount': annual_budget,
                'period': 'Annual Budget',
                'is_subtotal': is_subtotal,
                'is_total': False,
                'property_name': property_name,
                'metadata': json.dumps({
                    'source': 'variance_report',
                    'account_code': str(acct_code).strip() if acct_code else None,
                }),
            })

    logger.info(f"Variance Report (new): parsed {len(items)} items (period: {report_period})")
    return items


def _parse_variance_old_format(ws, property_name: str = None) -> List[Dict[str, Any]]:
    """
    Parse old-style variance report (pre-Oct 2025, summary format).

    Layout:
      Row 1: Title text
      Row 2: Section dividers ("Current Period" / "Year-To-Date")
      Row 3: Column headers (Actual, Budget, Variance, %, Explanation, Actual, Budget, Variance)
      Row 4+: Line items in col A, amounts in col B+
    """
    items = []

    # Try to extract period from sheet title or row 1
    report_period = None
    # Check sheet title (sometimes named "January 2025" etc.)
    if ws.title and re.match(r'\w+ \d{4}', ws.title):
        report_period = ws.title.strip()
    # Or from row 1 title text
    row1 = ws.cell(row=1, column=1).value
    if row1 and isinstance(row1, str):
        m = re.search(r'((?:January|February|March|April|May|June|July|August|'
                      r'September|October|November|December)\s+\d{4})', row1, re.IGNORECASE)
        if m:
            report_period = m.group(1)

    # Map headers from row 3
    # Old format: col B=Actual, C=Budget, D=Variance, E=%, F=Explanation,
    #             G=Actual(YTD), H=Budget(YTD), I=Variance(YTD)
    # But columns can shift — detect dynamically
    header_row = 3
    ptd_actual_col = None
    ptd_budget_col = None
    ptd_variance_col = None
    ptd_note_col = None
    ytd_actual_col = None
    ytd_budget_col = None
    ytd_variance_col = None

    actual_cols = []
    for col in range(1, 15):
        val = ws.cell(row=header_row, column=col).value
        if val and isinstance(val, str):
            v = val.strip().lower()
            if v == 'actual':
                actual_cols.append(col)
            elif v == 'budget' and not ptd_budget_col:
                ptd_budget_col = col
            elif v in ('variance', 'var') and not ptd_variance_col:
                ptd_variance_col = col
            elif 'explanation' in v or 'note' in v:
                ptd_note_col = col

    if len(actual_cols) >= 2:
        ptd_actual_col = actual_cols[0]
        ytd_actual_col = actual_cols[1]
        # Budget/variance for YTD are after the YTD actual
        for col in range(ytd_actual_col + 1, ytd_actual_col + 4):
            val = ws.cell(row=header_row, column=col).value
            if val and isinstance(val, str):
                v = val.strip().lower()
                if v == 'budget':
                    ytd_budget_col = col
                elif v in ('variance', 'var'):
                    ytd_variance_col = col
    elif len(actual_cols) == 1:
        ptd_actual_col = actual_cols[0]

    logger.info(f"Variance Report (old): PTD actual=col{ptd_actual_col}, "
                f"YTD actual=col{ytd_actual_col}, period={report_period}")

    current_category = 'revenue'
    current_subcategory = None

    for row_idx in range(4, ws.max_row + 1):
        line_name = ws.cell(row=row_idx, column=1).value
        if not line_name:
            continue
        if not isinstance(line_name, str):
            continue

        line_name_str = line_name.strip()
        if not line_name_str:
            continue

        line_lower = line_name_str.lower()
        is_total = line_lower.startswith('total') or line_lower.startswith('net ')

        # Update category tracking
        cat_key = line_lower.replace('total ', '').strip()
        if cat_key in _OS_CATEGORY_MAP:
            if not is_total:
                current_subcategory = line_name_str
            current_category = _OS_CATEGORY_MAP.get(line_lower, current_category)
        elif line_lower in _OS_CATEGORY_MAP:
            current_category = _OS_CATEGORY_MAP[line_lower]

        # Read amounts
        ptd_actual = _safe_float(ws.cell(row=row_idx, column=ptd_actual_col).value) if ptd_actual_col else None
        ptd_budget = _safe_float(ws.cell(row=row_idx, column=ptd_budget_col).value) if ptd_budget_col else None
        ptd_variance = _safe_float(ws.cell(row=row_idx, column=ptd_variance_col).value) if ptd_variance_col else None
        ytd_actual = _safe_float(ws.cell(row=row_idx, column=ytd_actual_col).value) if ytd_actual_col else None
        ytd_budget = _safe_float(ws.cell(row=row_idx, column=ytd_budget_col).value) if ytd_budget_col else None
        ytd_variance = _safe_float(ws.cell(row=row_idx, column=ytd_variance_col).value) if ytd_variance_col else None
        note = ws.cell(row=row_idx, column=ptd_note_col).value if ptd_note_col else None

        if ptd_actual is None and ytd_actual is None:
            continue

        is_subtotal = is_total

        if ptd_actual is not None:
            items.append({
                'category': current_category,
                'subcategory': current_subcategory,
                'line_item': line_name_str,
                'amount': ptd_actual,
                'period': f'{report_period} PTD' if report_period else 'PTD',
                'is_subtotal': is_subtotal,
                'is_total': False,
                'property_name': property_name,
                'metadata': json.dumps({
                    'source': 'variance_report',
                    'budget': ptd_budget,
                    'variance': ptd_variance,
                    'note': str(note).strip() if note else None,
                }),
            })

        if ytd_actual is not None:
            items.append({
                'category': current_category,
                'subcategory': current_subcategory,
                'line_item': line_name_str,
                'amount': ytd_actual,
                'period': f'{report_period} YTD' if report_period else 'YTD',
                'is_subtotal': is_subtotal,
                'is_total': False,
                'property_name': property_name,
                'metadata': json.dumps({
                    'source': 'variance_report',
                    'budget': ytd_budget,
                    'variance': ytd_variance,
                    'note': str(note).strip() if note else None,
                }),
            })

    logger.info(f"Variance Report (old): parsed {len(items)} items (period: {report_period})")
    return items


# ═══════════════════════════════════════════════════════════════════════
#  Top-level: Parse a full monthly report package
# ═══════════════════════════════════════════════════════════════════════

class MonthlyReportPackage:
    """
    Parsed results from a full Village Green monthly report package.

    Holds all extracted data, ready to be stored into the database tables.
    """
    def __init__(self):
        self.property_name: Optional[str] = None
        self.period: Optional[str] = None
        self.rent_roll_entries: List[Dict] = []       # → rent_roll_entries table
        self.operating_items: List[Dict] = []          # → operating_statement_items table
        self.financial_terms: List[Dict] = []          # → financial_terms table
        self.source_files: List[str] = []

    @property
    def summary(self) -> Dict[str, int]:
        return {
            'rent_roll_entries': len(self.rent_roll_entries),
            'operating_items': len(self.operating_items),
            'financial_terms': len(self.financial_terms),
            'source_files': len(self.source_files),
        }


def parse_monthly_report_package(
    excel_paths: List[str],
    property_name: str = None,
) -> MonthlyReportPackage:
    """
    Parse a complete monthly report package from one or more Excel files.

    Args:
        excel_paths: List of paths to the Excel files (Executive Summary,
                     12-Month Statement, Variance Report)
        property_name: Override property name (auto-detected if None)

    Returns:
        MonthlyReportPackage with all extracted data
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed — run: pip install openpyxl")
        return MonthlyReportPackage()

    pkg = MonthlyReportPackage()
    pkg.property_name = property_name

    for path in excel_paths:
        if not os.path.exists(path):
            logger.warning(f"File not found: {path}")
            continue

        fname = os.path.basename(path)
        pkg.source_files.append(fname)

        # Detect period from filename
        if not pkg.period:
            pkg.period = _parse_period_from_filename(fname)

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            logger.error(f"Failed to open {fname}: {e}")
            continue

        sheet_names_lower = {name.lower(): name for name in wb.sheetnames}

        # ── Executive Summary Report ──
        if 'exhibit 6 unit mix' in sheet_names_lower:
            # This is the Executive Summary Report workbook
            logger.info(f"Parsing Executive Summary Report: {fname}")

            if not pkg.property_name:
                pkg.property_name = _detect_property_name(wb[sheet_names_lower.get('executive summary', wb.sheetnames[0])])

            # Unit Mix → rent_roll_entries
            ws = wb[sheet_names_lower['exhibit 6 unit mix']]
            pkg.rent_roll_entries.extend(parse_unit_mix(ws))

            # Renewals → rent_roll_entries (individual units)
            if 'exhibit 4 renewals' in sheet_names_lower:
                ws = wb[sheet_names_lower['exhibit 4 renewals']]
                pkg.rent_roll_entries.extend(parse_renewals(ws))

            # NTV → rent_roll_entries
            if 'exhibit 3 ntv' in sheet_names_lower:
                ws = wb[sheet_names_lower['exhibit 3 ntv']]
                pkg.rent_roll_entries.extend(parse_ntv(ws))

            # Delinquency → financial_terms
            if 'exhibit 5 delinquency' in sheet_names_lower:
                ws = wb[sheet_names_lower['exhibit 5 delinquency']]
                pkg.financial_terms.extend(parse_delinquency(ws))

            # Occupancy Summary → financial_terms
            if 'exhibit 1 summary' in sheet_names_lower:
                ws = wb[sheet_names_lower['exhibit 1 summary']]
                pkg.financial_terms.extend(parse_occupancy_summary(ws))

            # Executive Summary sheet → top-level property metrics
            if 'executive summary' in sheet_names_lower:
                ws = wb[sheet_names_lower['executive summary']]
                pkg.financial_terms.extend(_parse_exec_summary_metrics(ws))

        # ── 12-Month Statement ──
        elif 'statement' in fname.lower() or (
            len(wb.sheetnames) == 1 and _is_12month_statement(wb.active)
        ):
            logger.info(f"Parsing 12-Month Statement: {fname}")
            ws = wb.active
            if not pkg.property_name:
                pkg.property_name = _detect_property_name(ws)
            pkg.operating_items.extend(
                parse_12month_statement(ws, property_name=pkg.property_name)
            )

        # ── Variance Report ──
        elif 'variance' in fname.lower() or (
            len(wb.sheetnames) == 1 and _is_variance_report(wb.active)
        ):
            logger.info(f"Parsing Variance Report: {fname}")
            ws = wb.active
            if not pkg.property_name:
                pkg.property_name = _detect_property_name(ws)
            pkg.operating_items.extend(
                parse_variance_report(ws, property_name=pkg.property_name)
            )

        wb.close()

    # ── Post-processing: compute occupancy from unit mix data ──
    # The Exhibit 1 Summary sheet often has stale/static occupancy values.
    # The unit mix has actual per-style occupied/vacant counts.
    _compute_occupancy_from_unit_mix(pkg)

    logger.info(
        f"Monthly report package complete: {pkg.summary} "
        f"(property: {pkg.property_name}, period: {pkg.period})"
    )
    return pkg


def _compute_occupancy_from_unit_mix(pkg: MonthlyReportPackage) -> None:
    """
    Replace stale Exhibit 1 occupancy with accurate occupancy computed
    from unit mix quantity/occupied/vacant counts.

    The Exhibit 1 Summary sheet in VG reports often carries a static
    occupancy figure that doesn't change month to month. The unit mix
    sheet has per-style occupied/vacant counts that reflect reality.
    """
    import json

    total_units = 0
    total_occupied = 0
    total_vacant = 0

    for entry in pkg.rent_roll_entries:
        raw_meta = entry.get('metadata')
        if not raw_meta:
            continue
        meta = json.loads(raw_meta) if isinstance(raw_meta, str) else raw_meta
        if meta.get('source') != 'unit_mix':
            continue
        qty = meta.get('quantity') or 0
        occ = meta.get('occupied') or 0
        vac = meta.get('vacant') or 0
        total_units += qty
        total_occupied += occ
        total_vacant += vac

    if total_units == 0:
        return  # No unit mix data to compute from

    computed_rate = round(total_occupied / total_units * 100, 1)
    computed_vacant = total_units - total_occupied

    logger.info(
        f"Computed occupancy from unit mix: {total_occupied}/{total_units} = "
        f"{computed_rate}% (vs Exhibit 1 static values)"
    )

    # Remove any existing Exhibit 1 occupancy_rate terms (they're stale)
    pkg.financial_terms = [
        t for t in pkg.financial_terms
        if t.get('term_type') != 'occupancy_rate'
    ]

    # Add computed occupancy
    pkg.financial_terms.append({
        'term_type': 'occupancy_rate',
        'term_label': 'Physical Occupancy (Computed from Unit Mix)',
        'value_raw': f'{computed_rate}%',
        'value_numeric': computed_rate,
    })

    # Add occupied/vacant counts
    pkg.financial_terms.append({
        'term_type': 'occupancy_count',
        'term_label': 'Occupied Units',
        'value_raw': str(total_occupied),
        'value_numeric': float(total_occupied),
    })
    pkg.financial_terms.append({
        'term_type': 'occupancy_count',
        'term_label': 'Vacant Units',
        'value_raw': str(computed_vacant),
        'value_numeric': float(computed_vacant),
    })

    # Also update total_units if not already present
    has_total_units = any(
        t.get('term_type') == 'total_units' for t in pkg.financial_terms
    )
    if not has_total_units:
        pkg.financial_terms.append({
            'term_type': 'total_units',
            'term_label': 'Total Units (from Unit Mix)',
            'value_raw': str(total_units),
            'value_numeric': float(total_units),
        })


def _is_12month_statement(ws) -> bool:
    """Detect if a worksheet looks like a 12-month P&L statement."""
    row2 = ws.cell(row=2, column=1).value
    if row2 and isinstance(row2, str) and 'statement' in row2.lower():
        return True
    # Check for multiple month columns in row 5
    month_count = 0
    for col in range(3, 20):
        val = ws.cell(row=5, column=col).value
        if val and isinstance(val, str) and any(m in val for m in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']):
            month_count += 1
    return month_count >= 6


def _is_variance_report(ws) -> bool:
    """Detect if a worksheet looks like a variance/budget report."""
    row2 = ws.cell(row=2, column=1).value
    if row2 and isinstance(row2, str) and ('budget' in row2.lower() or 'variance' in row2.lower()):
        return True
    # New format: PTD/YTD columns in row 5
    for col in range(3, 15):
        val = ws.cell(row=5, column=col).value
        if val and isinstance(val, str) and ('ptd' in val.lower() or 'ytd' in val.lower()):
            return True
    # Old format: "Variance Explanation" in row 1, headers in row 3
    row1 = ws.cell(row=1, column=1).value
    if row1 and isinstance(row1, str) and 'variance' in row1.lower():
        return True
    # Old format: row 3 has "Actual", "Budget", "Variance" headers
    for col in range(2, 10):
        val = ws.cell(row=3, column=col).value
        if val and isinstance(val, str) and val.strip().lower() in ('actual', 'budget', 'variance'):
            return True
    return False
