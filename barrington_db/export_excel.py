"""
export_excel.py - Build the Barrington portfolio Excel deliverable from the DB.

Usage:
    python export_excel.py [--db DATA/barrington.db] [--out OUTPUT.xlsx] [--year 2026]

Tabs produced:
    Portfolio Summary | Portfolio Cash Flow | Capital by Property |
    Lease Rollover | one tab per property (monthly cash flow + capital + expirations)
All on-device; reads only the built SQLite DB.
"""
import argparse, datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from barrington_db.db import connect, MONTHS_2026
from barrington_db.portfolio import Portfolio
from barrington_db.property_module import PropertyModule
from barrington_db.report import DISPLAY_ORDER, MONTH_LABELS

NAVY = PatternFill("solid", fgColor="1F3864")
LIGHT = PatternFill("solid", fgColor="D9E1F2")
WHITE_BOLD = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
MONEY = '#,##0;(#,##0)'
PSF = '#,##0.00'
THIN = Side(style="thin", color="BFBFBF")
BOT = Border(bottom=THIN)
CAP_ITEMS = [("CAP_BUILDING", "Base Bldg/LL"), ("CAP_TI", "TI"),
             ("CAP_LC", "LC"), ("CAP_DEFERRED", "Deferred"), ("CAP_NONOP", "Non-Op")]

def f(v):
    return float(v) if v is not None else 0.0

def header_row(ws, r, labels, start=1):
    for i, lab in enumerate(labels):
        c = ws.cell(r, start + i, lab)
        c.fill = NAVY; c.font = WHITE_BOLD
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title(ws, text, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, text); c.font = Font(bold=True, size=14, color="1F3864")

def build(conn, out_path, year=2026):
    pf = Portfolio(conn)
    wb = Workbook(); wb.remove(wb.active)

    # ---- 1. Portfolio Summary -------------------------------------------------
    ws = wb.create_sheet("Portfolio Summary")
    title(ws, f"Barrington Office Portfolio - {year} Summary", 9)
    cols = ["Code", "Property", "Market", f"NOI {year}", "Occupied SF",
            "Tenants", "Exp. thru 2028", "TI+LC+Bldg Cap", "NOI / SF"]
    header_row(ws, 3, cols)
    r = 4
    tot_noi = tot_sf = tot_cap = 0
    for s in pf.summary_table():
        cap = sum(f(v) for k, v in s["capital_2026"].items()
                  if k in ("CAP_BUILDING", "CAP_TI", "CAP_LC"))
        noi, sf = f(s["noi_2026"]), f(s["occupied_sf"])
        tot_noi += noi; tot_sf += sf; tot_cap += cap
        vals = [s["code"], s["name"], s["market"], noi, sf, s["tenants"],
                s["expiring_thru_2028"], cap, (noi / sf if sf else 0)]
        for i, v in enumerate(vals):
            ws.cell(r, i + 1, v)
        r += 1
    for i, v in enumerate(["TOTAL", "", "", tot_noi, tot_sf, "", "", tot_cap,
                           (tot_noi / tot_sf if tot_sf else 0)]):
        c = ws.cell(r, i + 1, v); c.font = BOLD; c.fill = LIGHT
    for row in range(4, r + 1):
        for col in (4, 5, 8):
            ws.cell(row, col).number_format = MONEY
        ws.cell(row, 9).number_format = PSF
    ws.freeze_panes = "A4"; widths(ws, [9, 30, 13, 14, 13, 9, 14, 16, 10])

    # ---- 2. Portfolio Cash Flow (annual) -------------------------------------
    ws = wb.create_sheet("Portfolio Cash Flow")
    title(ws, f"Portfolio Annual Cash Flow - {year}", 3)
    header_row(ws, 3, ["Line Item", "Section", f"{year} Total"])
    r = 4
    for row in pf.annual_cashflow(year):
        ws.cell(r, 1, row["label"]); ws.cell(r, 2, row["section"])
        c = ws.cell(r, 3, f(row["tot"])); c.number_format = MONEY
        if row["line_item_code"] in ("NOI",):
            for col in (1, 2, 3): ws.cell(r, col).font = BOLD
        r += 1
    ws.freeze_panes = "A4"; widths(ws, [34, 16, 16])

    # ---- 3. Capital by Property ----------------------------------------------
    ws = wb.create_sheet("Capital by Property")
    title(ws, f"Forecasted Capital by Property - {year}", len(CAP_ITEMS) + 2)
    header_row(ws, 3, ["Property"] + [lbl for _, lbl in CAP_ITEMS] + ["Total"])
    mat = {}
    for row in pf.capital_by_property(year):
        mat.setdefault(row["property_code"], {})[row["line_item_code"]] = row["tot"]
    r = 4; coltot = {c: 0 for c, _ in CAP_ITEMS}; grand = 0
    for code in pf.members:
        row = mat.get(code, {})
        ws.cell(r, 1, code)
        rowtot = 0
        for i, (c, _) in enumerate(CAP_ITEMS):
            v = f(row.get(c)); rowtot += v; coltot[c] += v
            cell = ws.cell(r, 2 + i, v); cell.number_format = MONEY
        grand += rowtot
        ws.cell(r, 2 + len(CAP_ITEMS), rowtot).number_format = MONEY
        r += 1
    ws.cell(r, 1, "TOTAL").font = BOLD
    for i, (c, _) in enumerate(CAP_ITEMS):
        cell = ws.cell(r, 2 + i, coltot[c]); cell.font = BOLD; cell.fill = LIGHT; cell.number_format = MONEY
    tc = ws.cell(r, 2 + len(CAP_ITEMS), grand); tc.font = BOLD; tc.fill = LIGHT; tc.number_format = MONEY
    ws.freeze_panes = "A4"; widths(ws, [12] + [14] * (len(CAP_ITEMS) + 1))

    # ---- 4. Lease Rollover ----------------------------------------------------
    ws = wb.create_sheet("Lease Rollover")
    title(ws, "Lease Rollover Through 2028 (drives forward TI/LC/Base-Bldg)", 4)
    header_row(ws, 3, ["Year", "Leases", "Expiring SF", "Expiring Rent"])
    r = 4
    for row in pf.rollover_by_year():
        ws.cell(r, 1, row["yr"]); ws.cell(r, 2, row["leases"])
        ws.cell(r, 3, f(row["sf"])).number_format = MONEY
        ws.cell(r, 4, f(row["rent"])).number_format = MONEY
        r += 1
    ws.freeze_panes = "A4"; widths(ws, [8, 10, 16, 18])

    # ---- 5..N Per-property tabs ----------------------------------------------
    for code in pf.members:
        pm = PropertyModule(conn, code)
        safe = code[:28]
        ws = wb.create_sheet(safe)
        title(ws, f"{pm.name} ({code}) - {year} Cash Flow  (values in $)", 14)
        header_row(ws, 3, ["Line Item"] + MONTH_LABELS + ["TOTAL"])
        monthly = pm.monthly(year)
        r = 4
        for licode, label in DISPLAY_ORDER:
            if licode not in monthly:
                continue
            ws.cell(r, 1, label.strip())
            vals = [f(monthly[licode].get(pid)) for pid in MONTHS_2026]
            for i, v in enumerate(vals):
                ws.cell(r, 2 + i, v).number_format = MONEY
            tc = ws.cell(r, 14, sum(vals)); tc.number_format = MONEY
            if licode == "NOI":
                for col in range(1, 15): ws.cell(r, col).font = BOLD
            r += 1
        # capital detail
        r += 1; ws.cell(r, 1, "Capital Detail").font = BOLD; r += 1
        for licode, label, tot in pm.capital_detail(year):
            ws.cell(r, 1, label); ws.cell(r, 2, f(tot)).number_format = MONEY; r += 1
        # expirations
        r += 1; ws.cell(r, 1, "Leases Expiring Through 2028").font = BOLD; r += 1
        header_row(ws, r, ["Tenant", "Area SF", "Lease To", "Annual Rent", "Rent PSF"]); r += 1
        for e in pm.expirations():
            ws.cell(r, 1, e["tenant_name"]); ws.cell(r, 2, f(e["area_sf"])).number_format = MONEY
            ws.cell(r, 3, e["lease_to"]); ws.cell(r, 4, f(e["annual_rent"])).number_format = MONEY
            ws.cell(r, 5, f(e["annual_rent_psf"])).number_format = PSF; r += 1
        ws.freeze_panes = "B4"; widths(ws, [30] + [11] * 12 + [13])

    wb.save(out_path)
    return out_path


# ─────────────────────────────────────────────────────────────────────────
# "Consolidated Cash Flow" style — a single sheet matching the client's
# layout: sections (Revenue, OpEx, NOI, Debt Service, Building Capital,
# Leasing Capital, Cash Flow, Loan Funding) with one row per property across
# the 12 months + Full Year, in $000s. Corporate Center I & II are combined.
# ─────────────────────────────────────────────────────────────────────────
_CONS_PROPS = [
    ("100/150 South Wacker", ["WACKER"]),
    ("O'Hare Plaza I", ["OHP1"]),
    ("O'Hare Plaza II", ["OHP2"]),
    ("Drake Oak Brook", ["DRAKE"]),
    ("Corporate Center I and II", ["CCI", "CCII"]),
    ("Combined Center", ["COMBINED"]),
    ("One Northbrook Place", ["ONB"]),
    ("JTM MKE LLC", ["JTM"]),
    ("MB MKE LLC", ["MB"]),
]
_CONS_SECTIONS = [
    ("revenue", "I.  Revenue", "Total Portfolio Revenue"),
    ("opex", "II.  Operating Expenses", "Total Portfolio Expenses"),
    ("noi", "III.  Net Operating Income", "Portfolio NOI"),
    ("debt", "IV.  Debt Service", "Portfolio Debt Service"),
    ("building", "V.  Building Capital", "Portfolio Building Capital"),
    ("leasing", "VI.  Leasing Capital", "Portfolio Leasing Capital"),
    ("cashflow", "VII.  Cash Flow", "Portfolio Cash Flow"),
    ("loan", "VIII.  Loan Funding Offset", "Portfolio Loan Funding"),
]


def _consolidated_data(conn, year):
    """Aggregate cash_flow_fact into per-property, per-section monthly arrays
    (12 months + Full Year). Returns {property_label: {section: [13 values]}}."""
    rows = conn.execute("""
        SELECT property_code, line_item_code, period_id, SUM(amount) amt
        FROM cash_flow_fact f JOIN period p USING(period_id)
        WHERE p.year = ? GROUP BY property_code, line_item_code, period_id
    """, (year,)).fetchall()
    piv = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for pc, code, pid, amt in rows:
        piv[pc][code][pid] += amt or 0.0

    def cm(codes, prop_codes, period):
        return sum(piv[pc][c].get(period, 0.0) for pc in prop_codes for c in codes)

    data = {}
    for label, pcs in _CONS_PROPS:
        # OHP I/II only have a combined CAP_SUBTOTAL (leasing-related); others
        # have the Building/TI/LC detail. Use detail when present.
        detail = any(c in piv[pc] for pc in pcs
                     for c in ('CAP_BUILDING', 'CAP_TI', 'CAP_LC', 'CAP_DEFERRED', 'CAP_NONOP'))

        def arr(codes):
            a = [cm(codes, pcs, p) for p in MONTHS_2026]
            a.append(sum(a))
            return a
        noi = arr({'NOI'})
        debt = arr({'INTEREST', 'PRINCIPAL'})
        bldg = arr({'CAP_BUILDING'})
        leas = arr({'CAP_TI', 'CAP_LC', 'CAP_DEFERRED', 'CAP_NONOP'}) if detail else arr({'CAP_SUBTOTAL'})
        cf = [noi[i] + debt[i] + bldg[i] + leas[i] for i in range(13)]
        loan = arr({'CONTRIBUTION'})
        data[label] = {
            'revenue': arr({'REVENUE', 'REVENUE_ADJ'}),
            'opex': arr({'OPEX', 'TAX_PAYMENT', 'INSURANCE_PAYMENT'}),
            'noi': noi, 'debt': debt, 'building': bldg, 'leasing': leas,
            'cashflow': cf, 'loan': loan,
            'net': [cf[i] + loan[i] for i in range(13)],
        }
    return data


def build_consolidated(conn, out_path, year=2026, title="Portfolio"):
    """Client 'Consolidated Cash Flow' workbook (single sheet, $000s)."""
    d = _consolidated_data(conn, year)
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year} Budget"
    ws.cell(1, 8, title).font = Font(bold=True, size=13)
    ws.cell(2, 8, "Consolidated Cash Flow").font = Font(bold=True, size=11)
    ws.cell(3, 8, f"{year} Budgets ($000s)").font = Font(italic=True)
    for i, m in enumerate(MONTH_LABELS):
        c = ws.cell(6, 3 + i, m); c.font = WHITE_BOLD; c.fill = NAVY
        c.alignment = Alignment(horizontal="center")
    c = ws.cell(6, 15, f"Full Year {year}"); c.font = WHITE_BOLD; c.fill = NAVY
    c.alignment = Alignment(horizontal="center")

    def put(r, idx, v, bold=False, fill=False):
        col = 15 if idx == 12 else 3 + idx
        show = round(v / 1000) if (v and abs(v) >= 500) else (0 if idx == 12 else None)
        cell = ws.cell(r, col, show)
        cell.number_format = MONEY
        if bold:
            cell.font = BOLD
        if fill:
            cell.fill = LIGHT
        return cell

    r = 8
    for key, roman, totlabel in _CONS_SECTIONS:
        ws.cell(r, 1, roman).font = BOLD
        r += 1
        tot = [0.0] * 13
        for label, _ in _CONS_PROPS:
            ws.cell(r, 1, label)
            vals = d[label][key]
            for i in range(13):
                put(r, i, vals[i])
                tot[i] += vals[i]
            r += 1
        r += 1
        ws.cell(r, 1, totlabel).font = BOLD
        for i in range(13):
            put(r, i, tot[i], bold=True, fill=True)
        r += 2

    net = [0.0] * 13
    for label, _ in _CONS_PROPS:
        for i in range(13):
            net[i] += d[label]['net'][i]
    ws.cell(r, 1, "PORTFOLIO NET CASH FLOW").font = Font(bold=True, size=12)
    for i in range(13):
        put(r, i, net[i], bold=True, fill=True)

    ws.column_dimensions['A'].width = 28
    for i in range(3, 16):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.freeze_panes = "C7"
    wb.save(out_path)
    return out_path


# Selectable output styles: name -> builder(conn, out_path, year[, title])
STYLES = {
    "summary": build,
    "consolidated": build_consolidated,
}


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", default=None)
    ap.add_argument("--out", default=f"Barrington_Portfolio_{datetime.date.today():%Y%m%d}.xlsx")
    ap.add_argument("--year", type=int, default=2026)
    ap.add_argument("--style", default="summary", choices=list(STYLES))
    a = ap.parse_args()
    conn = connect(a.db)
    if a.style == "consolidated":
        path = build_consolidated(conn, a.out, a.year, "Portfolio")
    else:
        path = build(conn, a.out, a.year)
    print("Wrote", path)
