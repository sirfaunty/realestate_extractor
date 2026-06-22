"""
extract_cashflow.py — Extract 2026 cash flow from source docs into the
canonical line-item / monthly fact schema.

Two extractors:
  * extract_drake_xlsx()  — clean Excel input sheet.
  * extract_pdf_cashflow()— generic layout parser for the PDF cash flows,
    matching canonical labels and pulling the 12 monthly columns.

Both emit a list of fact dicts:
    {property_code, period_id, line_item_code, scenario, amount,
     source_file, source_label}
"""
import re
import os
import sqlite3

import pdfplumber
from openpyxl import load_workbook

from .db import MONTHS_2026, scenario_for_period
from .reference import LINE_ITEM_ALIASES

NUM_RE = re.compile(r"\(?\$?-?[\d,]+(?:\.\d+)?\)?")


def _to_float(tok: str):
    """Parse an accounting-formatted token to float. Parens = negative."""
    if tok is None:
        return None
    s = str(tok).strip()
    if s in ("", "-", "$", "$-", "0"):
        return 0.0 if s == "0" else None
    neg = s.startswith("(") and s.endswith(")")
    s = s.replace("(", "").replace(")", "").replace("$", "").replace(",", "").strip()
    if s in ("", "-"):
        return None
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if neg else v


def _canon(label: str):
    """Map a raw source label to a canonical line_item_code, or None."""
    if not label:
        return None
    key = re.sub(r"\s+", " ", str(label).strip().lower())
    key = key.rstrip(":")
    return LINE_ITEM_ALIASES.get(key)


# ---------------------------------------------------------------------------
# Drake — Excel input sheet
# ---------------------------------------------------------------------------
def extract_drake_xlsx(path: str, property_code: str = "DRAKE"):
    wb = load_workbook(path, data_only=True)
    ws = wb["Sheet 1"]
    facts = []
    fname = os.path.basename(path)
    for row in ws.iter_rows(min_row=1, max_row=40, values_only=True):
        label = row[0]
        code = _canon(label)
        if not code:
            continue
        # Months are columns index 2..13 (C..N) == Jan..Dec.
        month_vals = row[2:14]
        for period_id, val in zip(MONTHS_2026, month_vals):
            amt = val if isinstance(val, (int, float)) else _to_float(val)
            if amt is None:
                continue
            facts.append({
                "property_code": property_code,
                "period_id": period_id,
                "line_item_code": code,
                "scenario": scenario_for_period(period_id),
                "amount": float(amt),
                "source_file": fname,
                "source_label": str(label).strip(),
            })
    return facts


# ---------------------------------------------------------------------------
# Generic PDF cash flow parser
# ---------------------------------------------------------------------------
def _extract_pdf_lines(path: str):
    """Return list of (text) lines from the cash flow PDF using layout text."""
    lines = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            txt = page.extract_text(layout=True) or ""
            lines.extend(txt.split("\n"))
    return lines


def _parse_money_row(line: str, expected: int = 12):
    """
    Given a line like 'Building Improvements   0   0   (380) ...',
    split into (label, [12 monthly floats]). Returns (label, vals) or None.
    Strategy: find all money-like tokens; the label is everything before the
    first token. We take the FIRST `expected` numeric values as Jan..Dec
    (TOTAL / budget / difference columns trail after and are ignored).
    """
    tokens = NUM_RE.findall(line)
    if len(tokens) < expected:
        return None
    # Label = text before first numeric token.
    first = NUM_RE.search(line)
    label = line[: first.start()].strip()
    if not label:
        return None
    vals = [_to_float(t) for t in tokens[:expected]]
    return label, vals


def extract_pdf_cashflow(path: str, property_code: str):
    """
    Parse a PDF cash flow into canonical monthly facts.
    Only rows whose label maps to a canonical code are captured, so noise
    (detail sub-lines, headers) is skipped. For properties with detailed
    capital schedules (JTM/MB/Wacker), the TOTAL roll-up rows are matched
    via aliases below.
    """
    facts = []
    fname = os.path.basename(path)
    seen = set()  # (code) -> only take first matching summary row per code
    for line in _extract_pdf_lines(path):
        norm = re.sub(r"\s+", " ", line).strip()
        if not norm:
            continue
        # Try canonical match on the leading label text.
        # Build candidate label by stripping a leading 'Less:'/'Add:' prefix.
        candidate = re.sub(r"^(less:|add:|deduct:|plus:)\s*", "", norm, flags=re.I)
        parsed = _parse_money_row(candidate)
        if not parsed:
            continue
        label, vals = parsed
        code = _canon(label)
        if not code:
            continue
        if code in seen:
            continue  # keep first occurrence (summary line)
        seen.add(code)
        for period_id, amt in zip(MONTHS_2026, vals):
            if amt is None:
                continue
            facts.append({
                "property_code": property_code,
                "period_id": period_id,
                "line_item_code": code,
                "scenario": scenario_for_period(period_id),
                "amount": float(amt),
                "source_file": fname,
                "source_label": label,
            })
    return facts


def load_facts(conn: sqlite3.Connection, facts):
    conn.executemany(
        "INSERT OR REPLACE INTO cash_flow_fact "
        "(property_code, period_id, line_item_code, scenario, amount, "
        " source_file, source_label) VALUES "
        "(:property_code, :period_id, :line_item_code, :scenario, :amount, "
        " :source_file, :source_label)",
        facts,
    )
    conn.commit()
    return len(facts)
