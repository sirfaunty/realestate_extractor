"""
Generate the Southtown / DHOS Co-Tenancy & Returns Model (Excel deliverable).

Live, formula-driven workbook mirroring the partner's five-sheet model:
  README · Dashboard · Returns (Return-on-Cost) · Scenarios (co-tenancy) · Tenant Roster

Blue cells are inputs you can change (tenant SF, Required? flags, TPC line items,
NOI, exit caps, scenario toggles); black cells are formulas that recompute. Driven
by returns_model.py; every headline number ties out to the partner's model.

Usage:
    python returns_xlsx.py --out Southtown_CoTenancy_and_Returns_Model.xlsx
"""
import argparse
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import returns_model as M

NAVY = "1B2A4A"
BLUE_INPUT = "0000FF"
CLOUD = "EEF3F8"
RED = "C8102E"
GREY = "606A78"

F_TITLE = Font(name="Arial", size=15, bold=True, color=NAVY)
F_SUB = Font(name="Arial", size=10, color="2E5A88")
F_H = Font(name="Arial", size=10, bold=True, color="FFFFFF")
F_BOLD = Font(name="Arial", size=10, bold=True, color=NAVY)
F_INPUT = Font(name="Arial", size=10, color=BLUE_INPUT)
F_CALC = Font(name="Arial", size=10, color="000000")
F_NOTE = Font(name="Arial", size=9, color=GREY)
FILL_H = PatternFill("solid", fgColor=NAVY)
FILL_CLOUD = PatternFill("solid", fgColor=CLOUD)
THIN = Side(style="thin", color="D5DEE8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT = "0.0%"
USD = '$#,##0;($#,##0);"-"'
SF = '#,##0;(#,##0);"-"'


def _hdr(ws, cell, text):
    ws[cell] = text
    ws[cell].font = F_H
    ws[cell].fill = FILL_H


def _title(ws, title, sub):
    ws["A1"] = "KRAUS-ANDERSON  |  Southtown Shopping Center"
    ws["A1"].font = F_SUB
    ws["A2"] = title
    ws["A2"].font = F_TITLE
    ws["A3"] = sub
    ws["A3"].font = F_SUB


def _roster_rows():
    return M.TENANT_ROSTER


def build(out_path):
    wb = Workbook()

    # ── Tenant Roster ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Tenant Roster"
    _title(ws, "Tenant Roster & Required-Tenant Classification",
           "Source: Southtown rent roll 5/31/2026. Blue = input.")
    hdr_row = 5
    for i, h in enumerate(["#", "Suite", "Occupant", "Total SF", "Lease Exp.", "Required?"], start=1):
        _hdr(ws, f"{get_column_letter(i)}{hdr_row}", h)
    r0 = hdr_row + 1
    for i, t in enumerate(_roster_rows()):
        r = r0 + i
        ws[f"A{r}"] = i + 1
        ws[f"B{r}"] = t["suite"]
        ws[f"C{r}"] = t["occupant"]
        ws[f"D{r}"] = t["sf"]; ws[f"D{r}"].font = F_INPUT; ws[f"D{r}"].number_format = SF
        ws[f"E{r}"] = t["exp"]
        ws[f"F{r}"] = "Y" if t["required"] else "N"; ws[f"F{r}"].font = F_INPUT
        ws[f"F{r}"].alignment = Alignment(horizontal="center")
    last = r0 + len(_roster_rows()) - 1
    qual_row = last + 2
    ws[f"C{qual_row}"] = "Qualifying SF if ALL Required Tenants open"
    ws[f"C{qual_row}"].font = F_BOLD
    ws[f"D{qual_row}"] = f"=SUMIF(F{r0}:F{last},\"Y\",D{r0}:D{last})"
    ws[f"D{qual_row}"].font = F_BOLD; ws[f"D{qual_row}"].number_format = SF
    ws[f"C{qual_row+1}"] = "Base LFA denominator (§1.6/1.7 stipulated)"
    ws[f"D{qual_row+1}"] = M.BASE_LFA; ws[f"D{qual_row+1}"].font = F_INPUT
    ws[f"D{qual_row+1}"].number_format = SF
    for col, w in zip("ABCDEF", [4, 10, 30, 12, 12, 11]):
        ws.column_dimensions[col].width = w
    # named cells for cross-sheet refs
    QUAL = f"'Tenant Roster'!$D${qual_row}"
    BASE = f"'Tenant Roster'!$D${qual_row+1}"
    kohl_row = r0 + 0    # Kohl's 1st floor is first roster entry
    gc_idx = next(i for i, t in enumerate(_roster_rows()) if t["occupant"] == "Guitar Center")
    KOHL = f"'Tenant Roster'!$D${r0 + 0}"
    GC = f"'Tenant Roster'!$D${r0 + gc_idx}"

    # ── Scenarios (co-tenancy engine) ───────────────────────────────────────
    sc = wb.create_sheet("Scenarios")
    _title(sc, "Co-Tenancy Scenarios — 65% Required-Tenant Test",
           "Each column is a scenario. Toggle blue cells; the test recomputes.")
    cols = [get_column_letter(2 + i) for i in range(len(M.SCENARIOS))]  # B..F
    sc["A5"] = "Driver / Line"; sc["A5"].font = F_BOLD
    for c, s in zip(cols, M.SCENARIOS):
        sc[f"{c}5"] = s["name"]; sc[f"{c}5"].font = F_BOLD
        sc[f"{c}5"].alignment = Alignment(horizontal="center", wrap_text=True)

    def rowlabel(r, text, bold=False):
        sc[f"A{r}"] = text
        sc[f"A{r}"].font = F_BOLD if bold else F_CALC

    rowlabel(6, "INPUTS / TOGGLES", True)
    rowlabel(7, "Kohl’s 1st floor status")
    rowlabel(8, "Guitar Center status")
    rowlabel(9, "Southtown Lanes redeveloped to non-retail")
    rowlabel(11, "NUMERATOR — qualifying open Required SF", True)
    rowlabel(12, "Full qualifying SF (all Required open)")
    rowlabel(13, "Less: Kohl’s 1st floor if vacant")
    rowlabel(14, "Less: Guitar Center if vacant")
    rowlabel(15, "Numerator = qualifying open SF")
    rowlabel(17, "DENOMINATOR — §3.11(ix) adjusted base", True)
    rowlabel(18, "Stipulated base LFA")
    rowlabel(19, "Less: Southtown Lanes non-retail carve-out")
    rowlabel(20, "Denominator = adjusted base")
    rowlabel(22, "CO-TENANCY TEST", True)
    rowlabel(23, "Occupancy % (numerator / denominator)")
    rowlabel(24, "65% required")
    rowlabel(25, "Cushion / (shortfall) vs. 65% (SF)")
    rowlabel(26, "PASS / FAIL")
    rowlabel(28, "RENT CONSEQUENCE", True)
    rowlabel(29, "Dick’s rent status")

    for c, s in zip(cols, M.SCENARIOS):
        kohl_vac = "Kohl’s — 1st floor" in s["vacate"]
        gc_vac = "Guitar Center" in s["vacate"]
        sc[f"{c}7"] = "Vacant" if kohl_vac else "Open"; sc[f"{c}7"].font = F_INPUT
        sc[f"{c}8"] = "Vacant" if gc_vac else "Open"; sc[f"{c}8"].font = F_INPUT
        sc[f"{c}9"] = "Yes" if s["lanes_redeveloped"] else "No"; sc[f"{c}9"].font = F_INPUT
        sc[f"{c}12"] = f"={QUAL}"
        sc[f"{c}13"] = f'=IF({c}7="Vacant",{KOHL},0)'
        sc[f"{c}14"] = f'=IF({c}8="Vacant",{GC},0)'
        sc[f"{c}15"] = f"={c}12-{c}13-{c}14"
        sc[f"{c}18"] = f"={BASE}"
        sc[f"{c}19"] = f'=IF({c}9="Yes",{M.LANES_SF},0)'
        sc[f"{c}20"] = f"={c}18-{c}19"
        sc[f"{c}23"] = f"={c}15/{c}20"; sc[f"{c}23"].number_format = PCT
        sc[f"{c}24"] = M.CO_TENANCY_THRESHOLD; sc[f"{c}24"].number_format = PCT
        sc[f"{c}24"].font = F_INPUT
        sc[f"{c}25"] = f"={c}15-{c}24*{c}20"; sc[f"{c}25"].number_format = SF
        sc[f"{c}26"] = f'=IF({c}23>={c}24,"PASS","FAIL")'
        sc[f"{c}29"] = f'=IF({c}26="PASS","Full Minimum Rent","SUBSTITUTE RENT — +18-mo termination right")'
        for rr in (12, 13, 14, 15, 18, 19, 20):
            sc[f"{c}{rr}"].number_format = SF
        sc[f"{c}26"].alignment = Alignment(horizontal="center")
    sc.column_dimensions["A"].width = 42
    for c in cols:
        sc.column_dimensions[c].width = 15

    # ── Dashboard ───────────────────────────────────────────────────────────
    db = wb.create_sheet("Dashboard")
    _title(db, "DHOS Co-Tenancy — Headline Pass/Fail",
           "65% Required-Tenant test across scenarios (pulls from Scenarios).")
    for i, h in enumerate(["Scenario", "Occupancy %", "Cushion/(Short) SF", "Result", "Rent impact"], 1):
        _hdr(db, f"{get_column_letter(i)}6", h)
    for i, (c, s) in enumerate(zip(cols, M.SCENARIOS)):
        r = 7 + i
        db[f"A{r}"] = s["name"]
        db[f"B{r}"] = f"=Scenarios!{c}23"; db[f"B{r}"].number_format = PCT
        db[f"C{r}"] = f"=Scenarios!{c}25"; db[f"C{r}"].number_format = SF
        db[f"D{r}"] = f"=Scenarios!{c}26"; db[f"D{r}"].alignment = Alignment(horizontal="center")
        db[f"E{r}"] = f"=Scenarios!{c}29"
    for col, w in zip("ABCDE", [26, 13, 18, 10, 40]):
        db.column_dimensions[col].width = w

    # ── Returns (Return-on-Cost) ────────────────────────────────────────────
    rt = wb.create_sheet("Returns")
    _title(rt, "DHOS Return-on-Cost",
           "USES verbatim from Brama’s native proforma. Blue = input.")
    for i, h in enumerate(["Use (verbatim line)", "Amount", "$/SF", "Cash?"], 1):
        _hdr(rt, f"{get_column_letter(i)}5", h)
    u0 = 6
    for i, (label, amt, cash) in enumerate(M.TPC_USES):
        r = u0 + i
        rt[f"A{r}"] = label
        rt[f"B{r}"] = amt; rt[f"B{r}"].font = F_INPUT; rt[f"B{r}"].number_format = USD
        rt[f"C{r}"] = f"=B{r}/$B${{lfa}}".replace("{lfa}", "")  # set after lfa known
        rt[f"D{r}"] = "Cash" if cash else "NON-CASH"; rt[f"D{r}"].alignment = Alignment(horizontal="center")
    ulast = u0 + len(M.TPC_USES) - 1
    tpc_row = ulast + 1
    rt[f"A{tpc_row}"] = "TOTAL PROJECT COST (TPC)"; rt[f"A{tpc_row}"].font = F_BOLD
    rt[f"B{tpc_row}"] = f"=SUM(B{u0}:B{ulast})"; rt[f"B{tpc_row}"].font = F_BOLD
    rt[f"B{tpc_row}"].number_format = USD

    key = tpc_row + 2
    rt[f"A{key}"] = "KEY FIGURES"; rt[f"A{key}"].font = F_BOLD
    rt[f"A{key+1}"] = "Building LFA (SF)"; rt[f"B{key+1}"] = M.BUILDING_LFA
    rt[f"B{key+1}"].font = F_INPUT; rt[f"B{key+1}"].number_format = SF
    rt[f"A{key+2}"] = "Stabilized NOI (no TIF)"; rt[f"B{key+2}"] = M.STABILIZED_NOI
    rt[f"B{key+2}"].font = F_INPUT; rt[f"B{key+2}"].number_format = USD
    rt[f"A{key+3}"] = "Land value (KA-owned, non-cash)"; rt[f"B{key+3}"] = M.LAND_VALUE
    rt[f"B{key+3}"].font = F_INPUT; rt[f"B{key+3}"].number_format = USD
    LFA = f"$B${key+1}"; NOI = f"$B${key+2}"; LAND = f"$B${key+3}"; TPC = f"$B${tpc_row}"
    # now set $/SF formulas for uses
    for i in range(len(M.TPC_USES)):
        r = u0 + i
        rt[f"C{r}"] = f"=B{r}/{LFA}"; rt[f"C{r}"].number_format = "0.00"
    rt[f"C{tpc_row}"] = f"=B{tpc_row}/{LFA}"; rt[f"C{tpc_row}"].number_format = "0.00"

    yv = key + 5
    rt[f"A{yv}"] = "TWO YIELD-ON-COST VIEWS"; rt[f"A{yv}"].font = F_BOLD
    for i, h in enumerate(["Basis", "Amount", "Yield on Cost", "Read"], 1):
        _hdr(rt, f"{get_column_letter(i)}{yv+1}", h)
    rt[f"A{yv+2}"] = "All-in TPC (incl. owned land)"
    rt[f"B{yv+2}"] = f"={TPC}"; rt[f"B{yv+2}"].number_format = USD
    rt[f"C{yv+2}"] = f"={NOI}/{TPC}"; rt[f"C{yv+2}"].number_format = PCT
    rt[f"D{yv+2}"] = "Brama headline"
    rt[f"A{yv+3}"] = "CASH BASIS (excl. owned land)"
    rt[f"B{yv+3}"] = f"={TPC}-{LAND}"; rt[f"B{yv+3}"].number_format = USD
    rt[f"C{yv+3}"] = f"={NOI}/({TPC}-{LAND})"; rt[f"C{yv+3}"].number_format = PCT
    rt[f"D{yv+3}"] = "True cash out the door"

    ev = yv + 5
    rt[f"A{ev}"] = "VALUE vs. BASIS — exit-cap sensitivity"; rt[f"A{ev}"].font = F_BOLD
    for i, h in enumerate(["Exit cap", "Value", "vs all-in TPC", "vs CASH basis"], 1):
        _hdr(rt, f"{get_column_letter(i)}{ev+1}", h)
    for i, cap in enumerate(M.EXIT_CAP_RANGE):
        r = ev + 2 + i
        rt[f"A{r}"] = cap; rt[f"A{r}"].font = F_INPUT; rt[f"A{r}"].number_format = PCT
        rt[f"B{r}"] = f"={NOI}/A{r}"; rt[f"B{r}"].number_format = USD
        rt[f"C{r}"] = f"=B{r}-{TPC}"; rt[f"C{r}"].number_format = USD
        rt[f"D{r}"] = f"=B{r}-({TPC}-{LAND})"; rt[f"D{r}"].number_format = USD
    for col, w in zip("ABCD", [42, 16, 15, 16]):
        rt.column_dimensions[col].width = w

    # ── README ──────────────────────────────────────────────────────────────
    rd = wb.create_sheet("README")
    _title(rd, "Co-Tenancy & Returns Model", "Companion to the DHOS Lease Strategic Review")
    notes = [
        "", "WHAT THIS MODEL DOES",
        "Tests the §1.6/§1.7 co-tenancy condition (65% of Required-Tenant LFA open) across",
        "vacancy scenarios, and computes DHOS development returns (Yield-on-Cost, exit-cap value).",
        "", "THE KEY MECHANIC",
        "Occupancy = qualifying open Required-Tenant SF ÷ §3.11(ix)-adjusted base LFA (281,637).",
        "Below 65%, Dick’s pays Substitute Rent and gains an 18-month termination right.",
        "", "HOW TO USE IT",
        "1. Tenant Roster: blue cells (SF, Required?) are inputs.",
        "2. Scenarios: toggle Kohl’s / Guitar Center / Southtown Lanes (blue cells).",
        "3. Dashboard: headline pass/fail. Returns: TPC, Yield-on-Cost, exit-cap value.",
        "", "COLOR KEY",
        "Blue text = input you can change.  Black text = formula.",
        "", "PROVENANCE",
        "Roster from Southtown rent roll 5/31/2026; USES verbatim from Brama’s native proforma.",
        "Every headline number ties out to the source model.",
    ]
    for i, line in enumerate(notes):
        rd[f"A{5+i}"] = line
        rd[f"A{5+i}"].font = F_BOLD if line.isupper() and line else F_NOTE
    rd.column_dimensions["A"].width = 95
    wb.move_sheet("README", -(len(wb.sheetnames) - 1))  # README first

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="Southtown_CoTenancy_and_Returns_Model.xlsx")
    args = ap.parse_args()
    p = build(args.out)
    print(f"Wrote {p}")
