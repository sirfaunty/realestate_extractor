"""
Extract the TPC USES from Brama's native proforma (local, on-device).

Reads the "Sources & Uses" sheet: the ten USES line items (labels + amounts in
column J, rows 36-45) and the Total Project Cost. Feeds the returns engine.

Usage:
    from extract_proforma import extract_uses
    uses, tpc = extract_uses("source_docs/returns/Brama_NATIVE_proforma_AUTHORITATIVE.xlsx")
"""
import os
import sys

_HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = os.path.join(_HERE, "source_docs", "returns",
                            "Brama_NATIVE_proforma_AUTHORITATIVE.xlsx")

# USES block in "Sources & Uses": labels in the row, amount in column J (10).
_USES_ROWS = range(36, 46)
_AMOUNT_COL = 10          # column J
_NON_CASH_LABELS = {"land"}   # land is the KA-owned, non-cash basis item


def extract_uses(xlsx_path=DEFAULT_XLSX):
    """Return (uses, tpc) where uses = [(label, amount, is_cash)]."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Sources & Uses"]
    uses = []
    for r in _USES_ROWS:
        amt = ws.cell(r, _AMOUNT_COL).value
        if not isinstance(amt, (int, float)):
            continue
        label = None
        for c in range(1, _AMOUNT_COL):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                label = v.strip()
                break
        if label is None:
            continue
        is_cash = label.strip().lower() not in _NON_CASH_LABELS
        uses.append((label, round(amt, 2), is_cash))
    wb.close()
    tpc = round(sum(a for _l, a, _c in uses))
    return uses, tpc


def available(xlsx_path=DEFAULT_XLSX):
    return os.path.exists(xlsx_path)


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    uses, tpc = extract_uses(path)
    print(f"Extracted {len(uses)} USES line items — TPC ${tpc:,}")
    for label, amt, cash in uses:
        print(f"  {'Cash    ' if cash else 'NON-CASH'}  ${amt:>14,.2f}  {label[:50]}")
